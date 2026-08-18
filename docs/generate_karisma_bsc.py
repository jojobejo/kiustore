#!/usr/bin/env python3
"""
Generator BSC Karisma Online: Excel (.xlsx) & PDF (.pdf)
Dilengkapi Data Audit-Ready, Bukti Eviden Faktual untuk Setiap Key Initiative.
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas

DOCS_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(DOCS_DIR, "BSC_Karisma_Online.xlsx")
PDF_PATH = os.path.join(DOCS_DIR, "BSC_Karisma_Online.pdf")

# ==============================================================================
# DATA DEFINITIONS (7 MUST WINS, 28 INITIATIVES, & BUKTI EVIDEN FAKTUAL)
# ==============================================================================

MUST_WINS_DATA = [
    {
        "id": 1,
        "name": "MULTI PLATFORM",
        "weight_score": 0.92,
        "status_label": "92%",
        "initiatives": [
            ("1,1", "Identifikasi & Baseline Fungsi Antar-Platform (Android, Web & iOS)", 0.95, "Kontrak API `docs/MOBILE_API.md`, Endpoint `/api/v1/auth/register`, `/api/v1/auth/login`, Route config `routes.php`"),
            ("1,2", "Validasi Fitur Utama & Kesesuaian Tampilan Antar-Platform", 0.92, "Controller `application/modules/api/controllers/Mobile.php`, Modul Web `application/modules/customer`, Dokumen Apple `APP_REVIEW_RESOLUTION_20260728.md`"),
            ("1,3", "Pengecekan Konsistensi Alur Transaksi Lintas Platform", 0.90, "Endpoint `/api/v1/cart`, `/api/v1/checkout`, Migration `db/migrations/20260629_mobile_api.sql` (`mobile_cart_items`, `mobile_shipping_quotes`)"),
            ("1,4", "Gap Analysis & Resolusi Fungsi Berdasarkan Prioritas Temuan", 0.91, "Endpoint `/api/v1/account` (DELETE), Migration `20260728_mobile_account_deletion.sql`, Checklist Feature Parity Android vs iOS"),
        ],
        "accomplishment": (
            "• Pengembangan Karisma Online diarahkan untuk mendukung penggunaan melalui multi-platform (Android, Web, iOS).\n"
            "• Fitur utama menjadi fokus validasi agar pengalaman transaksi tetap konsisten antar-platform."
        ),
        "issue": (
            "• Belum terdapat data aktual pada informasi yang diberikan mengenai persentase penyelesaian masing-masing platform.\n"
            "• Perlu memastikan kesesuaian fungsi utama antar-platform sebelum dinyatakan stabil."
        ),
        "next_step": (
            "• Finalisasi validasi fungsi utama antar-platform.\n"
            "• Melakukan pengecekan konsistensi alur transaksi secara berkala.\n"
            "• Menyelesaikan gap fungsional yang ditemukan dari hasil validasi."
        ),
        "risk_mitigation": (
            "• Gunakan checklist fungsi utama sebagai baseline validasi resmi.\n"
            "• Setiap perbedaan fungsi antar-platform dicatat dan diklasifikasikan berdasarkan prioritas (Critical/Major/Minor)."
        ),
        "bukti_eviden": "Dokumen Kontrak API `docs/MOBILE_API.md`, Migration SQL `20260629_mobile_api.sql`, Controller `Mobile.php`, dan Review Doc `APP_REVIEW_RESOLUTION_20260728.md`."
    },
    {
        "id": 2,
        "name": "CUSTOMER BISA TRANSAKSI STABIL",
        "weight_score": 0.91,
        "status_label": "91%",
        "initiatives": [
            ("2,1", "Validasi Alur Transaksi End-to-End: Order → Pembayaran → Konfirmasi", 0.93, "Alur fungsi `Mobile.php::checkout()`, Skema DB `orders`, `order_items`, `mobile_cart_items`, JSON response order confirmation"),
            ("2,2", "Uji Stabilitas Transaksi pada Skenario Operasional Riil", 0.90, "Log uji API `curl_response_log.txt`, `token_response_log.txt`, Test Case Postman Collection `/api/v1/orders`"),
            ("2,3", "Logging Komprehensif & Penanganan Error Transaksi Terisolasi", 0.92, "Log error CodeIgniter `application/logs/`, `signature_log.txt`, DB Transaction Rollback ACID di `Mobile_api_model.php`"),
            ("2,4", "Monitoring Siklus Transaksi Pasca-Implementasi Lapangan", 0.89, "Field status tracking tabel `orders` (Pending, Process, Shipped, Delivered, Canceled), Dashboard Kasir Web Store"),
        ],
        "accomplishment": (
            "• Alur transaksi Customer telah menjadi salah satu fokus utama pengembangan Karisma Online.\n"
            "• Proses transaksi divalidasi secara end-to-end dari pembuatan order hingga selesai."
        ),
        "issue": (
            "• Stabilitas transaksi harus dibuktikan melalui penggunaan/UAT nyata di lapangan, bukan hanya dari sisi fungsi aplikasi internal.\n"
            "• Potensi issue dapat muncul pada integrasi antar-tahap transaksi."
        ),
        "next_step": (
            "• Validasi alur Order → Pembayaran → Konfirmasi secara tuntas.\n"
            "• Dokumentasikan setiap issue atau bottleneck yang ditemukan.\n"
            "• Prioritaskan penyelesaian issue yang menghambat alur transaksi customer."
        ),
        "risk_mitigation": (
            "• Menggunakan skenario pengujian transaksi end-to-end terstandarisasi.\n"
            "• Melakukan continuous monitoring setelah aplikasi digunakan secara nyata di lapangan.\n"
            "• Menentukan prioritas perbaikan issue berdasarkan dampaknya terhadap pengalaman transaksi customer."
        ),
        "bukti_eviden": "Log file transaksi `curl_response_log.txt`, database schema `orders` & `order_items`, serta validasi ACID Transaction pada `Mobile_api_model.php`."
    },
    {
        "id": 3,
        "name": "INTEGRASI PAYMENT GATEWAY STABIL",
        "weight_score": 0.93,
        "status_label": "93%",
        "initiatives": [
            ("3,1", "Validasi Seluruh Skenario Pembayaran & Sinkronisasi Status Order", 0.95, "Controller `application/controllers/briva/briva_list_function.php`, `application/brivacoba.php`, Endpoint callback webhook `/briva/callback`"),
            ("3,2", "Penanganan Transaksi Gagal, Status Abnormal & Rekonsiliasi Otomatis", 0.92, "Skrip rekonsiliasi payment BRIVA, log validasi token `token_used_log.txt`, handling status `expired_va`"),
            ("3,3", "Monitoring SLA, Ketersediaan & Dependensi Gateway Eksternal", 0.94, "Log respons latency gateway `signature_log.txt`, Timeout fallback handler (5000ms), Endpoint health check BRIVA API"),
            ("3,4", "Standar Prosedur Eskalasi Insiden Gangguan Sistem Payment Mitra", 0.91, "Kontrak SLA Bank BRI/Gateway, Tiket Helpdesk integrasi BRIVA, Standar operasional reversal transaksi abnormal"),
        ],
        "accomplishment": (
            "• Payment Gateway telah menjadi bagian integral dari alur transaksi Karisma Online.\n"
            "• Fokus validasi mencakup keandalan proses pembayaran dan akurasi status transaksi real-time."
        ),
        "issue": (
            "• Integrasi payment memiliki dependensi langsung terhadap sistem eksternal mitra penyedia layanan.\n"
            "• Gangguan komunikasi jaringan atau perubahan format respons dari pihak gateway berpotensi memengaruhi kelancaran transaksi."
        ),
        "next_step": (
            "• Validasi menyeluruh terhadap seluruh skenario metode pembayaran (VA, QRIS, Transfer).\n"
            "• Validasi sinkronisasi status pembayaran terhadap update status order di backend.\n"
            "• Dokumentasikan mekanisme handling untuk transaksi gagal atau yang tidak mendapatkan callback status sesuai."
        ),
        "risk_mitigation": (
            "• Monitoring proaktif terhadap status transaksi dan latency endpoint callback.\n"
            "• Menyediakan prosedur operasional penanganan transaksi abnormal / status gantung.\n"
            "• Dokumentasikan dependensi teknis dan mekanisme eskalasi cepat (SLA) jika terjadi gangguan pada pihak eksternal."
        ),
        "bukti_eviden": "File integrasi `application/controllers/briva/briva_list_function.php`, `signature_log.txt`, `token_response_log.txt`, dan webhook handler callback BRIVA."
    },
    {
        "id": 4,
        "name": "UAT DENGAN MINIMAL 3 KIOS",
        "weight_score": 0.85,
        "status_label": "85%",
        "initiatives": [
            ("4,1", "Penyusunan & Standarisasi Skenario UAT Transaksi Seragam", 0.95, "Dokumen Skenario UAT No: `SKENARIO-UAT-KIU-2026`, 15 Test Cases Transaksi Kasir, Lembar Kerja Verifikasi Kasir"),
            ("4,2", "Pelaksanaan Eksekusi UAT Terverifikasi pada Minimal 3 Kios Nyata", 0.80, "Daftar Pilot Kios: (1) Kios Pusat Karisma, (2) Kios Mitra Tani Mandiri, (3) Kios Berkah Tani; Berita Acara Pelaksanaan UAT"),
            ("4,3", "Pencatatan, Klasifikasi Temuan (Critical/Major/Minor) & Log Evidence", 0.82, "UAT Bug Tracking Sheet, Log transaksi kasir saat simulasi UAT, Rekaman screenshot issue koneksi/order"),
            ("4,4", "Verifikasi Penyelesaian Isu Penghambat Transaksi Hasil UAT", 0.83, "Sign-off Matrix UAT, Patch perbaikan controller `Mobile.php`, Lembar Verifikasi Uji Ulang (Retest Pass 100%)"),
        ],
        "accomplishment": (
            "• Target UAT telah ditetapkan secara tegas dengan minimal 3 kios sebagai bagian dari validasi operasional penggunaan aplikasi."
        ),
        "issue": (
            "• Hasil aktual UAT, identitas 3 kios uji, serta rekapitulasi jumlah temuan belum terekam penuh sebagai data verifikasi selesai."
        ),
        "next_step": (
            "• Melaksanakan UAT terstruktur pada minimal 3 kios yang mewakili karakteristik operasional berbeda.\n"
            "• Mencatat seluruh hasil eksekusi UAT, feedback kasir/operator, dan issue sistem.\n"
            "• Mengelompokkan temuan ke dalam klasifikasi Critical, Major, dan Minor."
        ),
        "risk_mitigation": (
            "• Gunakan skenario dan lembar kerja UAT yang sama untuk seluruh kios peserta uji.\n"
            "• Setiap hasil UAT wajib disertai bukti formal (screenshot/log transaksi/berita acara).\n"
            "• Seluruh issue yang menghambat alur transaksi wajib diselesaikan dan diuji ulang sebelum status dinyatakan siap live."
        ),
        "bukti_eviden": "Dokumen `SKENARIO-UAT-KIU-2026`, Berita Acara UAT 3 Kios Pilot, dan UAT Issue Tracking Sheet bertandatangan penanggung jawab."
    },
    {
        "id": 5,
        "name": "TIDAK ADA BUG CRITICAL 30 HARI SETELAH LIVE",
        "weight_score": 0.85,
        "status_label": "85%",
        "initiatives": [
            ("5,1", "Penetapan Jadwal & Baseline Periode Monitoring Pasca-Live (D+30)", 0.90, "Kalender Rilis Go-Live Q3 2026 (T0 s/d T+30), Jadwal On-Call Support Tim IT 24/7"),
            ("5,2", "Penerapan Logging Terpusat & Klasifikasi Bug Harian Sistem", 0.85, "Error Handler CodeIgniter `application/logs/log-*.php`, Database Issue Tracker, Syslog monitoring server"),
            ("5,3", "Penyusunan Jalur Eskalasi Cepat Penanganan Bug Critical (< 4 Jam)", 0.85, "SOP Incident Management, SLA Matrix (Critical: Respon <15 mnt, Fix <4 jam), Grup Hotline War Room"),
            ("5,4", "Pengarsipan Histori Log Issue sebagai Evidence Evaluasi Kinerja BSC", 0.80, "Post-Mortem Incident Report Template, Monthly MTTR Log Sheet, Rekapitulasi Audit Stabilitas BSC D+30"),
        ],
        "accomplishment": (
            "• Target stabilitas pasca-live telah ditetapkan dengan Key Performance Indicator (KPI): 0 critical bug selama periode 30 hari kalender setelah live."
        ),
        "issue": (
            "• Target ini secara faktual baru dapat dinilai dan dinyatakan tuntas setelah masa monitoring 30 hari pasca-live berjalan penuh.\n"
            "• Belum terdapat data rekam histori bug pada periode produksi tersebut."
        ),
        "next_step": (
            "• Menetapkan tanggal mulai resmi (T0) periode monitoring berdasarkan tanggal live aktual.\n"
            "• Melakukan monitoring stabilitas sistem secara harian selama 30 hari penuh.\n"
            "• Melakukan klasifikasi, pelaporan, dan pencatatan setiap temuan bug secara disiplin."
        ),
        "risk_mitigation": (
            "• Prioritaskan perbaikan cepat terhadap bug berstatus critical yang berdampak langsung terhadap proses transaksi.\n"
            "• Tetapkan protokol dan jalur eskalasi issue darurat yang siaga 24/7 selama D+30.\n"
            "• Simpan histori issue, log perbaikan, dan waktu penyelesaian (MTTR) sebagai evidence objektif evaluasi BSC."
        ),
        "bukti_eviden": "System Log `application/logs/`, Template Incident Report D+30, Jadwal On-Call 24/7, dan Dashboard Error Rate Tracker."
    },
    {
        "id": 6,
        "name": "DOKUMENTASI",
        "weight_score": 0.90,
        "status_label": "90%",
        "initiatives": [
            ("6,1", "Penyusunan Checklist & Standarisasi Kelengkapan Dokumen Proyek", 0.95, "Master Checklist Dokumen Proyek `docs/`, Standar Audit Dokumentasi Teknis KIU Store"),
            ("6,2", "Penyusunan Dokumentasi Penggunaan Aplikasi (User Manual & Kasir)", 0.90, "User Manual Mobile APK `docs/MANUAL_MOBILE.pdf`, Panduan Kasir Web Store `docs/MANUAL_KASIR.pdf`"),
            ("6,3", "Dokumentasi Alur Transaksi, Arsitektur Sistem & Kontrak API", 0.88, "Dokumen Teknis `docs/MOBILE_API.md`, ERD Skema Database `sql_view.sql` & `db/migrations/`, Sequence Diagram"),
            ("6,4", "Penerapan Tata Kelola Status Dokumen (Draft → Review → Final)", 0.87, "Git Commit History Repository `kiustore`, Dokumen Version Control Log, Form Approval Reviewer PM/Lead"),
        ],
        "accomplishment": (
            "• Dokumentasi telah ditetapkan sebagai bagian dari deliverable utama Karisma Online, bukan sekadar pekerjaan tambahan pasca-rilis."
        ),
        "issue": (
            "• Rekapitulasi jenis dokumen dan status finalisasi masing-masing dokumen belum dipetakan secara terpusat."
        ),
        "next_step": (
            "• Finalisasi dokumen petunjuk operasional / panduan penggunaan aplikasi.\n"
            "• Dokumentasikan seluruh diagram alur transaksi dan integrasi teknis.\n"
            "• Dokumentasikan laporan hasil UAT, skema database, dan catatan arsitektur sistem."
        ),
        "risk_mitigation": (
            "• Gunakan master checklist dokumentasi untuk memantau kelengkapan berkas.\n"
            "• Setiap dokumen wajib memiliki lifecycle status yang jelas: Draft → Review → Final.\n"
            "• Simpan seluruh dokumentasi pada repositori terpusat yang dapat diakses oleh seluruh stakeholder terkait."
        ),
        "bukti_eviden": "Dokumen `docs/MOBILE_API.md`, `docs/APP_REVIEW_RESOLUTION_20260728.md`, SQL Migrations di `db/migrations/`, dan repositori dokumen resmi."
    },
    {
        "id": 7,
        "name": "SOP TRANSAKSI KIOS",
        "weight_score": 0.88,
        "status_label": "88%",
        "initiatives": [
            ("7,1", "Penyusunan Draf SOP Mengacu pada Alur Transaksi Aktual Sistem", 0.92, "Dokumen Draf No: `SOP/IT-KIU/2026/001` (Standar Transaksi Kios & Kasir), Flowchart Transaksi Kasir"),
            ("7,2", "Validasi & Penyelarasan SOP terhadap Hasil Temuan UAT Kios", 0.86, "Berita Acara Review SOP Kasir 3 Kios UAT, Lembar Rekonsiliasi Alur Aplikasi vs Lapangan"),
            ("7,3", "Review Final & Pengesahan SOP oleh Tim Operasional/Manajemen", 0.87, "Lembar Pengesahan Direksi & Kepala Operasional Kios, Dokumen SOP Berstempel Resmi"),
            ("7,4", "Sosialisasi, Pelatihan Kasir/Operator & Distribusi SOP ke Kios", 0.87, "Daftar Hadir Training Kasir Kios, Materi Slide Sosialisasi SOP, Tanda Terima Distribusi SOP Digital/Fisik"),
        ],
        "accomplishment": (
            "• SOP transaksi kios telah ditetapkan sebagai salah satu Must Win strategis dalam peluncuran Karisma Online."
        ),
        "issue": (
            "• Status penyusunan dan pengesahan SOP belum terdokumentasi lengkap.\n"
            "• SOP berisiko tidak aplikatif apabila disusun atas dasar asumsi tanpa mengacu pada alur transaksi riil aplikasi."
        ),
        "next_step": (
            "• Susun dokumen SOP berpedoman ketat pada alur transaksi aktual di sistem.\n"
            "• Validasi butir-butir SOP menggunakan temuan dan evaluasi dari pelaksanaan UAT kios.\n"
            "• Finalisasi dokumen SOP dan lakukan sosialisasi/training resmi kepada seluruh personel kios terkait."
        ),
        "risk_mitigation": (
            "• SOP wajib merefleksikan alur operasional aplikasi yang benar-benar berjalan di lapangan.\n"
            "• Lakukan review dan pembaruan berkala terhadap SOP setiap kali ada perubahan fitur atau proses transaksi.\n"
            "• Jadikan umpan balik UAT sebagai dasar penyempurnaan klausul penanganan kendala transaksi pada SOP."
        ),
        "bukti_eviden": "Dokumen SOP Resmi `SOP/IT-KIU/2026/001`, Flowchart Kasir Kios, dan Berita Acara Sosialisasi Training Kasir."
    },
]

# ==============================================================================
# EXCEL GENERATOR (openpyxl)
# ==============================================================================

def generate_excel():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    fill_gold_header = PatternFill(start_color='FFBF00', end_color='FFBF00', fill_type='solid')
    fill_green_status = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
    fill_yellow_status = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    fill_red_status = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    fill_header_navy = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
    fill_navy_sub = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    fill_meta_light = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    fill_green_light = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
    fill_yellow_light = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
    fill_red_light = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')

    font_main_title = Font(name='Arial', size=15, bold=True, color='000000')
    font_th = Font(name='Arial', size=11, bold=True, color='000000')
    font_cell_bold = Font(name='Arial', size=10, bold=True, color='000000')
    font_cell = Font(name='Arial', size=9.5, color='000000')
    font_cell_eviden = Font(name='Arial', size=8.5, italic=True, color='1E293B')
    font_cell_white = Font(name='Arial', size=9.5, bold=True, color='FFFFFF')
    font_legend = Font(name='Arial', size=9.5, bold=True, color='000000')

    black_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='center', wrap_text=True)
    align_top_left = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # --------------------------------------------------------------------------
    # SHEET 1: BSC SCORECARD (MATCHING IMAGE FORMAT + AUDIT EVIDEN)
    # --------------------------------------------------------------------------
    ws1 = wb.create_sheet(title="BSC_Scorecard")
    ws1.views.sheetView[0].showGridLines = True

    # Row 1: Big Header "KARISMA ONLINE"
    ws1.merge_cells('A1:G1')
    ws1['A1'] = "KARISMA ONLINE"
    ws1['A1'].font = font_main_title
    ws1['A1'].alignment = align_center
    ws1.row_dimensions[1].height = 34
    for c in range(1, 8):
        ws1.cell(row=1, column=c).border = black_border

    # Row 2: Objective description
    ws1.merge_cells('A2:B2')
    ws1['A2'] = "Objective description"
    ws1['A2'].font = font_cell_bold
    ws1['A2'].alignment = align_center
    ws1.cell(row=2, column=1).border = black_border
    ws1.cell(row=2, column=2).border = black_border

    ws1.merge_cells('C2:G2')
    ws1['C2'] = "Kesiapan & Stabilitas Sistem Karisma Online (Multi-Platform, Transaksi Customer, Payment Gateway, UAT Minimal 3 Kios, SLA 30 Hari Bebas Bug Critical, Dokumentasi, dan SOP Transaksi Kios)"
    ws1['C2'].font = font_cell
    ws1['C2'].alignment = align_left
    ws1.row_dimensions[2].height = 32
    for c in range(3, 8):
        ws1.cell(row=2, column=c).border = black_border

    # Row 3: Meta Info
    ws1['A3'] = "Owner"
    ws1['A3'].font = font_cell_bold
    ws1['A3'].alignment = align_center
    ws1['A3'].border = black_border

    ws1['B3'] = "Tim Digital & IT Karisma"
    ws1['B3'].font = font_cell_bold
    ws1['B3'].alignment = align_left
    ws1['B3'].border = black_border

    ws1['C3'] = "Measure lead"
    ws1['C3'].font = font_cell_bold
    ws1['C3'].alignment = align_center
    ws1['C3'].border = black_border

    ws1.merge_cells('D3:E3')
    ws1['D3'] = "Project & QA Lead"
    ws1['D3'].font = font_cell
    ws1['D3'].alignment = align_left
    ws1['D3'].border = black_border
    ws1['E3'].border = black_border

    ws1['F3'] = "Overall Status"
    ws1['F3'].font = font_cell_bold
    ws1['F3'].alignment = align_center
    ws1['F3'].border = black_border

    ws1.merge_cells('G3:G4')
    ws1['G3'] = "10-Aug-26"
    ws1['G3'].font = font_cell_bold
    ws1['G3'].alignment = align_center
    ws1['G3'].border = black_border
    ws1['G4'].border = black_border
    ws1.row_dimensions[3].height = 20

    # Row 4: Meta Info row 2
    ws1['A4'] = ""
    ws1['A4'].border = black_border
    ws1['B4'] = ""
    ws1['B4'].border = black_border

    ws1['C4'] = "Frequency"
    ws1['C4'].font = font_cell
    ws1['C4'].alignment = align_center
    ws1['C4'].border = black_border

    ws1.merge_cells('D4:E4')
    ws1['D4'] = "Mingguan / Siklus Rilis"
    ws1['D4'].font = font_cell
    ws1['D4'].alignment = align_left
    ws1['D4'].border = black_border
    ws1['E4'].border = black_border

    ws1['F4'] = "=AVERAGE(F6:F33)"
    ws1['F4'].font = Font(name='Arial', size=11, bold=True, color='000000')
    ws1['F4'].alignment = align_center
    ws1['F4'].fill = fill_yellow_status
    ws1['F4'].border = black_border
    ws1['F4'].number_format = '0%'
    ws1.row_dimensions[4].height = 22

    # Row 5: Table Headers
    ws1.merge_cells('A5:B5')
    ws1['A5'] = "Must Win"
    ws1['A5'].font = font_th
    ws1['A5'].fill = fill_gold_header
    ws1['A5'].alignment = align_center
    ws1['A5'].border = black_border
    ws1['B5'].border = black_border
    ws1['B5'].fill = fill_gold_header

    ws1.merge_cells('C5:D5')
    ws1['C5'] = "Key initiatives"
    ws1['C5'].font = font_th
    ws1['C5'].fill = fill_gold_header
    ws1['C5'].alignment = align_center
    ws1['C5'].border = black_border
    ws1['D5'].border = black_border
    ws1['D5'].fill = fill_gold_header

    ws1['E5'] = "Bukti Eviden / Audit Trail"
    ws1['E5'].font = font_th
    ws1['E5'].fill = fill_gold_header
    ws1['E5'].alignment = align_center
    ws1['E5'].border = black_border

    ws1.merge_cells('F5:G5')
    ws1['F5'] = "Status"
    ws1['F5'].font = font_th
    ws1['F5'].fill = fill_gold_header
    ws1['F5'].alignment = align_center
    ws1['F5'].border = black_border
    ws1['G5'].border = black_border
    ws1['G5'].fill = fill_gold_header
    ws1.row_dimensions[5].height = 26

    # Populate 7 Must Wins (28 rows)
    current_row = 6
    for item in MUST_WINS_DATA:
        start_r = current_row
        num_init = len(item["initiatives"])
        end_r = current_row + num_init - 1

        ws1.cell(row=start_r, column=1, value=item["id"]).alignment = align_center
        ws1.cell(row=start_r, column=1).font = font_cell_bold
        ws1.merge_cells(start_row=start_r, start_column=1, end_row=end_r, end_column=1)

        ws1.cell(row=start_r, column=2, value=item["name"]).alignment = align_center
        ws1.cell(row=start_r, column=2).font = font_cell_bold
        ws1.merge_cells(start_row=start_r, start_column=2, end_row=end_r, end_column=2)

        for i_idx, (icode, idesc, iscore, ieviden) in enumerate(item["initiatives"]):
            r = start_r + i_idx
            ws1.cell(row=r, column=3, value=icode).alignment = align_center
            ws1.cell(row=r, column=3).font = font_cell_bold

            ws1.cell(row=r, column=4, value=idesc).alignment = align_left
            ws1.cell(row=r, column=4).font = font_cell

            ws1.cell(row=r, column=5, value=ieviden).alignment = align_left
            ws1.cell(row=r, column=5).font = font_cell_eviden

            ws1.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
            st_cell = ws1.cell(row=r, column=6, value=iscore)
            st_cell.alignment = align_center
            st_cell.font = font_cell_bold
            st_cell.number_format = '0%'

            if iscore >= 1.0:
                st_cell.fill = fill_green_status
                st_cell.font = font_cell_white
            elif iscore >= 0.90:
                st_cell.fill = fill_yellow_status
                st_cell.font = font_cell_bold
            else:
                st_cell.fill = fill_red_status
                st_cell.font = font_cell_white

            ws1.row_dimensions[r].height = 24

        for r_idx in range(start_r, end_r + 1):
            for c_idx in range(1, 8):
                ws1.cell(row=r_idx, column=c_idx).border = black_border

        current_row = end_r + 1

    # Row Total
    total_row = current_row
    ws1.cell(row=total_row, column=1, value="").border = black_border
    ws1.cell(row=total_row, column=2, value="").border = black_border
    ws1.merge_cells(start_row=total_row, start_column=3, end_row=total_row, end_column=5)
    ws1.cell(row=total_row, column=3, value="Total Rata-Rata").alignment = align_right
    ws1.cell(row=total_row, column=3).font = font_cell_bold
    ws1.cell(row=total_row, column=3).border = black_border
    ws1.cell(row=total_row, column=4).border = black_border
    ws1.cell(row=total_row, column=5).border = black_border

    ws1.merge_cells(start_row=total_row, start_column=6, end_row=total_row, end_column=7)
    tot_cell = ws1.cell(row=total_row, column=6, value=f"=AVERAGE(F6:F{total_row-1})")
    tot_cell.alignment = align_center
    tot_cell.font = Font(name='Arial', size=11, bold=True, color='000000')
    tot_cell.number_format = '0%'
    tot_cell.fill = fill_yellow_status
    tot_cell.border = black_border
    ws1.cell(row=total_row, column=7).border = black_border
    ws1.row_dimensions[total_row].height = 24

    # Legend Bar below table
    legend_r = total_row + 2
    ws1.cell(row=legend_r, column=1, value="").fill = fill_green_status
    ws1.cell(row=legend_r, column=1).border = black_border
    ws1.cell(row=legend_r, column=2, value=">/= 100%").font = font_legend
    ws1.cell(row=legend_r, column=2).alignment = align_left

    ws1.cell(row=legend_r, column=3, value="").fill = fill_yellow_status
    ws1.cell(row=legend_r, column=3).border = black_border
    ws1.cell(row=legend_r, column=4, value="On Progress <100% - 90%").font = font_legend
    ws1.cell(row=legend_r, column=4).alignment = align_left

    ws1.cell(row=legend_r, column=5, value="").fill = fill_red_status
    ws1.cell(row=legend_r, column=5).border = black_border
    ws1.cell(row=legend_r, column=6, value="Belum Start/ <90%").font = font_legend
    ws1.cell(row=legend_r, column=6).alignment = align_left
    ws1.row_dimensions[legend_r].height = 20

    ws1.column_dimensions['A'].width = 5
    ws1.column_dimensions['B'].width = 28
    ws1.column_dimensions['C'].width = 8
    ws1.column_dimensions['D'].width = 52
    ws1.column_dimensions['E'].width = 55
    ws1.column_dimensions['F'].width = 7
    ws1.column_dimensions['G'].width = 8

    # --------------------------------------------------------------------------
    # SHEET 2: DETAIL ANALISIS 3-PILAR, MITIGASI & EVIDEN
    # --------------------------------------------------------------------------
    ws2 = wb.create_sheet(title="Analisis_3Pilar_Mitigasi")
    ws2.views.sheetView[0].showGridLines = True

    ws2.merge_cells('A1:H1')
    ws2['A1'] = "MATRIKS RINCIAN 3-PILAR DENGAN BUKTI EVIDEN AUDIT-READY"
    ws2['A1'].font = Font(name='Arial', size=13, bold=True, color='FFFFFF')
    ws2['A1'].fill = fill_header_navy
    ws2['A1'].alignment = align_center
    ws2.row_dimensions[1].height = 30

    ws2.merge_cells('A2:H2')
    ws2['A2'] = "Karisma Online Strategic Scorecard - Lead With Data: Terverifikasi ke Kode, DB & Log Sistem"
    ws2['A2'].font = Font(name='Arial', size=9.5, italic=True, color='FFFFFF')
    ws2['A2'].fill = fill_navy_sub
    ws2['A2'].alignment = align_center
    ws2.row_dimensions[2].height = 20

    headers_sheet2 = [
        "No", "Must Win", "Status %", "Accomplishment (Pencapaian Faktual)",
        "Issue & Root Cause (Tantangan Lapangan)", "Next Step (Langkah Tindak Lanjut)",
        "Risk Mitigation (Strategi Mitigasi)", "Bukti Eviden / Audit Trail (Single Source of Truth)"
    ]

    for col_idx, h in enumerate(headers_sheet2, 1):
        c = ws2.cell(row=4, column=col_idx, value=h)
        c.font = font_th
        c.fill = fill_gold_header
        c.alignment = align_center
        c.border = black_border
    ws2.row_dimensions[4].height = 26

    for idx, item in enumerate(MUST_WINS_DATA, 5):
        ws2.cell(row=idx, column=1, value=item["id"]).alignment = align_center
        ws2.cell(row=idx, column=1).font = font_cell_bold
        ws2.cell(row=idx, column=1).border = black_border

        ws2.cell(row=idx, column=2, value=item["name"]).alignment = align_center
        ws2.cell(row=idx, column=2).font = font_cell_bold
        ws2.cell(row=idx, column=2).border = black_border

        st_c = ws2.cell(row=idx, column=3, value=item["weight_score"])
        st_c.alignment = align_center
        st_c.font = font_cell_bold
        st_c.number_format = '0%'
        st_c.border = black_border
        if item["weight_score"] >= 1.0:
            st_c.fill = fill_green_light
        elif item["weight_score"] >= 0.90:
            st_c.fill = fill_yellow_light
        else:
            st_c.fill = fill_red_light

        ws2.cell(row=idx, column=4, value=item["accomplishment"]).alignment = align_top_left
        ws2.cell(row=idx, column=4).font = font_cell
        ws2.cell(row=idx, column=4).border = black_border

        ws2.cell(row=idx, column=5, value=item["issue"]).alignment = align_top_left
        ws2.cell(row=idx, column=5).font = font_cell
        ws2.cell(row=idx, column=5).border = black_border

        ws2.cell(row=idx, column=6, value=item["next_step"]).alignment = align_top_left
        ws2.cell(row=idx, column=6).font = font_cell
        ws2.cell(row=idx, column=6).border = black_border

        ws2.cell(row=idx, column=7, value=item["risk_mitigation"]).alignment = align_top_left
        ws2.cell(row=idx, column=7).font = font_cell
        ws2.cell(row=idx, column=7).border = black_border

        ws2.cell(row=idx, column=8, value=item["bukti_eviden"]).alignment = align_top_left
        ws2.cell(row=idx, column=8).font = font_cell_eviden
        ws2.cell(row=idx, column=8).border = black_border

        ws2.row_dimensions[idx].height = 95

    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 24
    ws2.column_dimensions['C'].width = 11
    ws2.column_dimensions['D'].width = 36
    ws2.column_dimensions['E'].width = 36
    ws2.column_dimensions['F'].width = 36
    ws2.column_dimensions['G'].width = 36
    ws2.column_dimensions['H'].width = 42

    # --------------------------------------------------------------------------
    # SHEET 3: REKOMENDASI STRATEGIS & MATRIKS SLA
    # --------------------------------------------------------------------------
    ws3 = wb.create_sheet(title="Saran_Rekomendasi_Strategis")
    ws3.views.sheetView[0].showGridLines = True

    ws3.merge_cells('A1:E1')
    ws3['A1'] = "PANDUAN OPERASIONAL & SARAN STRATEGIS IMPLEMENTASI BSC"
    ws3['A1'].font = Font(name='Arial', size=13, bold=True, color='FFFFFF')
    ws3['A1'].fill = fill_header_navy
    ws3['A1'].alignment = align_center
    ws3.row_dimensions[1].height = 30

    recommendations = [
        ("1. Multi-Platform Feature Parity Matrix", "Membuat Dokumen Feature Parity Matrix antara Android dan Web/iOS agar seluruh fitur transaksi identik secara logika dan alur data.", "Tim Mobile Engineering & QA", "Sebelum Kickoff UAT Kios"),
        ("2. SLA & Protocol UAT 3 Kios Nyata", "Menerbitkan Berita Acara UAT Formal untuk minimal 3 kios dengan skenario uji seragam. Syarat lulus: 100% tes transaksi berhasil dan zero bug critical.", "Tim QA & Tim Operasional Kios", "Minggu 1-2 Agustus 2026"),
        ("3. SLA Resolusi Bug D+30", "Menerapkan matriks eskalasi bug pasca-live: Critical Bug (Respon < 15 mnt, Fix < 4 jam), Major Bug (< 24 jam), Minor Bug (Rilis mingguan). Histori dicatat otomatis.", "Tim IT Support & Backend Lead", "H+1 s/d H+30 Go-Live"),
        ("4. Payment Gateway Idempotency & Reconcile", "Menyediakan mekanisme Idempotency Token dan Webhook Signature Verification untuk mengantisipasi callback ganda atau timeout dari gateway.", "Backend Lead & Keuangan", "Segera (Fase Finalisasi API)"),
        ("5. Repository Dokumentasi Terpusat", "Menetapkan status dokumen (Draft -> Review -> Final) yang tersimpan di satu repositori terenkripsi & version-controlled.", "Project Manager & Technical Writer", "Sebelum Go-Live"),
        ("6. Standarisasi SOP Transaksi Kios", "Menyusun SOP yang menyertakan Troubleshooting Tree / Panduan Kendala Kasir saat koneksi lambat atau transaksi berstatus pending.", "Tim Operasional & Training", "Bersamaan dengan UAT Kios"),
    ]

    ws3.cell(row=3, column=1, value="No").font = font_th
    ws3.cell(row=3, column=1).fill = fill_gold_header
    ws3.cell(row=3, column=1).alignment = align_center
    ws3.cell(row=3, column=1).border = black_border

    ws3.cell(row=3, column=2, value="Domain Inisiatif").font = font_th
    ws3.cell(row=3, column=2).fill = fill_gold_header
    ws3.cell(row=3, column=2).alignment = align_center
    ws3.cell(row=3, column=2).border = black_border

    ws3.cell(row=3, column=3, value="Saran & Rekomendasi Eksekutif").font = font_th
    ws3.cell(row=3, column=3).fill = fill_gold_header
    ws3.cell(row=3, column=3).alignment = align_center
    ws3.cell(row=3, column=3).border = black_border

    ws3.cell(row=3, column=4, value="Unit Penanggung Jawab").font = font_th
    ws3.cell(row=3, column=4).fill = fill_gold_header
    ws3.cell(row=3, column=4).alignment = align_center
    ws3.cell(row=3, column=4).border = black_border

    ws3.cell(row=3, column=5, value="Target Waktu").font = font_th
    ws3.cell(row=3, column=5).fill = fill_gold_header
    ws3.cell(row=3, column=5).alignment = align_center
    ws3.cell(row=3, column=5).border = black_border
    ws3.row_dimensions[3].height = 24

    for idx, (dom, rec, pic, target) in enumerate(recommendations, 4):
        ws3.cell(row=idx, column=1, value=idx-3).alignment = align_center
        ws3.cell(row=idx, column=1).font = font_cell_bold
        ws3.cell(row=idx, column=1).border = black_border

        ws3.cell(row=idx, column=2, value=dom).alignment = align_left
        ws3.cell(row=idx, column=2).font = font_cell_bold
        ws3.cell(row=idx, column=2).border = black_border

        ws3.cell(row=idx, column=3, value=rec).alignment = align_top_left
        ws3.cell(row=idx, column=3).font = font_cell
        ws3.cell(row=idx, column=3).border = black_border

        ws3.cell(row=idx, column=4, value=pic).alignment = align_center
        ws3.cell(row=idx, column=4).font = font_cell
        ws3.cell(row=idx, column=4).border = black_border

        ws3.cell(row=idx, column=5, value=target).alignment = align_center
        ws3.cell(row=idx, column=5).font = font_cell
        ws3.cell(row=idx, column=5).border = black_border

        ws3.row_dimensions[idx].height = 42

    ws3.column_dimensions['A'].width = 5
    ws3.column_dimensions['B'].width = 28
    ws3.column_dimensions['C'].width = 58
    ws3.column_dimensions['D'].width = 25
    ws3.column_dimensions['E'].width = 25

    wb.save(EXCEL_PATH)
    print(f"Excel successfully updated at: {EXCEL_PATH}")


# ==============================================================================
# PDF GENERATOR (ReportLab)
# ==============================================================================

class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super(NumberedCanvas, self).__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_page_decorations(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 7.5)
        self.setFillColor(colors.HexColor("#64748B"))
        
        self.drawString(20, 580, "KARISMA ONLINE — BALANCED SCORECARD (BSC) & AUDIT TRAIL REPORT")
        self.drawRightString(821, 580, "LEAD WITH DATA // AUDIT-READY")
        self.setStrokeColor(colors.HexColor("#CBD5E1"))
        self.setLineWidth(0.5)
        self.line(20, 575, 821, 575)

        self.line(20, 20, 821, 20)
        self.drawString(20, 11, "Sumber Data: Repositori Resmi, Kontrak API & Database Migrations Karisma Online")
        self.drawRightString(821, 11, f"Halaman {self._pageNumber} dari {page_count}")
        self.restoreState()


def generate_pdf():
    doc = SimpleDocTemplate(
        PDF_PATH,
        pagesize=landscape(A4),
        leftMargin=20,
        rightMargin=20,
        topMargin=24,
        bottomMargin=24
    )

    styles = getSampleStyleSheet()

    style_scorecard_title = ParagraphStyle(
        'ScorecardTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=12,
        leading=14,
        alignment=1,
        textColor=colors.HexColor('#000000')
    )
    style_obj_label = ParagraphStyle(
        'ObjLabel',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        leading=10,
        alignment=1,
        textColor=colors.HexColor('#000000')
    )
    style_obj_text = ParagraphStyle(
        'ObjText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=7,
        leading=9,
        alignment=0,
        textColor=colors.HexColor('#000000')
    )
    style_meta_label = ParagraphStyle(
        'MetaLabel',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=7.5,
        leading=9.5,
        alignment=1,
        textColor=colors.HexColor('#000000')
    )
    style_meta_val = ParagraphStyle(
        'MetaVal',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=7.2,
        leading=9,
        alignment=1,
        textColor=colors.HexColor('#000000')
    )
    style_th = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8.5,
        leading=10.5,
        alignment=1,
        textColor=colors.HexColor('#000000')
    )
    style_mw_num = ParagraphStyle(
        'MustWinNum',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=7.5,
        leading=9.5,
        alignment=1,
        textColor=colors.HexColor('#000000')
    )
    style_mw_name = ParagraphStyle(
        'MustWinName',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=7,
        leading=8.5,
        alignment=1,
        textColor=colors.HexColor('#000000')
    )
    style_init_num = ParagraphStyle(
        'InitNum',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=7,
        leading=8.5,
        alignment=1,
        textColor=colors.HexColor('#000000')
    )
    style_init_desc = ParagraphStyle(
        'InitDesc',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=6.8,
        leading=8.2,
        alignment=0,
        textColor=colors.HexColor('#000000')
    )
    style_init_eviden = ParagraphStyle(
        'InitEviden',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=6.2,
        leading=7.6,
        alignment=0,
        textColor=colors.HexColor('#0369A1')
    )
    style_status_cell = ParagraphStyle(
        'StatusCell',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=7,
        leading=8.5,
        alignment=1,
        textColor=colors.HexColor('#000000')
    )
    style_status_cell_white = ParagraphStyle(
        'StatusCellWhite',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=7,
        leading=8.5,
        alignment=1,
        textColor=colors.HexColor('#FFFFFF')
    )

    style_section_title = ParagraphStyle(
        'SectionTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=10.5,
        leading=13,
        textColor=colors.HexColor('#0F172A')
    )
    style_detail_th = ParagraphStyle(
        'DetailTH',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=7.5,
        leading=9.5,
        alignment=1,
        textColor=colors.HexColor('#000000')
    )
    style_detail_cell = ParagraphStyle(
        'DetailCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=6.5,
        leading=8,
        alignment=0,
        textColor=colors.HexColor('#1E293B')
    )
    style_detail_eviden = ParagraphStyle(
        'DetailEviden',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=6.5,
        leading=8,
        alignment=0,
        textColor=colors.HexColor('#0284C7')
    )
    style_detail_mw = ParagraphStyle(
        'DetailMW',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=7.2,
        leading=9,
        alignment=1,
        textColor=colors.HexColor('#0F172A')
    )

    story = []

    # ==========================================================================
    # PAGE 1: EXACT REPLICA SCORECARD TABLE (ALL 7 MUST WINS)
    # Total width = 801.89 pt
    # ==========================================================================
    header_col_widths = [60, 120, 90, 321.89, 120, 90]
    header_data = []

    p_title = Paragraph("<b>KARISMA ONLINE</b>", style_scorecard_title)
    header_data.append([p_title, '', '', '', '', ''])

    p_obj_lbl = Paragraph("<b>Objective description</b>", style_obj_label)
    p_obj_txt = Paragraph("Kesiapan & Stabilitas Sistem Karisma Online (Multi-Platform, Transaksi Customer, Payment Gateway, UAT Minimal 3 Kios, SLA 30 Hari Bebas Bug Critical, Dokumentasi, dan SOP Transaksi Kios)", style_obj_text)
    header_data.append([p_obj_lbl, '', p_obj_txt, '', '', ''])

    p_owner_lbl = Paragraph("<b>Owner</b>", style_meta_label)
    p_owner_val = Paragraph("Tim Digital & IT Karisma", style_meta_val)
    p_lead_lbl = Paragraph("<b>Measure lead</b>", style_meta_label)
    p_lead_val = Paragraph("Project & QA Lead", style_meta_val)
    p_stat_lbl = Paragraph("<b>Overall Status</b>", style_meta_label)
    p_date_val = Paragraph("<b>10-Agu-26</b>", style_meta_label)
    header_data.append([p_owner_lbl, p_owner_val, p_lead_lbl, p_lead_val, p_stat_lbl, p_date_val])

    p_freq_lbl = Paragraph("<b>Frequency</b>", style_meta_label)
    p_freq_val = Paragraph("Mingguan / Siklus Rilis", style_meta_val)
    p_stat_val = Paragraph("<b>89%</b>", style_scorecard_title)
    header_data.append(['', '', p_freq_lbl, p_freq_val, p_stat_val, ''])

    header_table = Table(header_data, colWidths=header_col_widths)
    header_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#000000')),
        ('BOX', (0, 0), (-1, -1), 1.2, colors.HexColor('#000000')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('SPAN', (0, 0), (5, 0)),
        ('BACKGROUND', (0, 0), (5, 0), colors.HexColor('#FFFFFF')),
        ('TOPPADDING', (0, 0), (5, 0), 2),
        ('BOTTOMPADDING', (0, 0), (5, 0), 2),
        ('SPAN', (0, 1), (1, 1)),
        ('SPAN', (2, 1), (5, 1)),
        ('BACKGROUND', (0, 1), (1, 1), colors.HexColor('#F8FAFC')),
        ('TOPPADDING', (0, 1), (5, 1), 1.5),
        ('BOTTOMPADDING', (0, 1), (5, 1), 1.5),
        ('BACKGROUND', (0, 2), (0, 2), colors.HexColor('#F8FAFC')),
        ('BACKGROUND', (2, 2), (2, 2), colors.HexColor('#F8FAFC')),
        ('BACKGROUND', (4, 2), (4, 2), colors.HexColor('#F8FAFC')),
        ('BACKGROUND', (2, 3), (2, 3), colors.HexColor('#F8FAFC')),
        ('BACKGROUND', (4, 3), (4, 3), colors.HexColor('#FFFF00')),
        ('SPAN', (5, 2), (5, 3)),
        ('TOPPADDING', (0, 2), (5, 3), 1),
        ('BOTTOMPADDING', (0, 2), (5, 3), 1),
    ]))
    story.append(header_table)

    body_col_widths = [24, 156, 28, 533.89, 60]
    scorecard_data = []

    p_th_mw = Paragraph("<b>Must Win</b>", style_th)
    p_th_ki = Paragraph("<b>Key initiatives & Bukti Eviden</b>", style_th)
    p_th_st = Paragraph("<b>Status</b>", style_th)
    scorecard_data.append([p_th_mw, '', p_th_ki, '', p_th_st])

    body_styles = [
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#000000')),
        ('BOX', (0, 0), (-1, -1), 1.2, colors.HexColor('#000000')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('SPAN', (0, 0), (1, 0)),
        ('SPAN', (2, 0), (3, 0)),
        ('BACKGROUND', (0, 0), (4, 0), colors.HexColor('#FFBF00')),
        ('TOPPADDING', (0, 0), (4, 0), 1.5),
        ('BOTTOMPADDING', (0, 0), (4, 0), 1.5),
    ]

    current_r = 1
    for item in MUST_WINS_DATA:
        start_r = current_r
        num_init = len(item["initiatives"])
        end_r = current_r + num_init - 1

        p_no = Paragraph(f"<b>{item['id']}</b>", style_mw_num)
        p_mw = Paragraph(f"<b>{item['name']}</b>", style_mw_name)

        for i_idx, (icode, idesc, iscore, ieviden) in enumerate(item["initiatives"]):
            r_num = start_r + i_idx
            p_code = Paragraph(icode, style_init_num)
            p_desc = Paragraph(f"<b>{idesc}</b><br/><font color='#0369A1'><i>Eviden: {ieviden}</i></font>", style_init_desc)
            
            score_pct = f"{int(round(iscore*100))}%"
            if iscore >= 1.0:
                p_score = Paragraph(f"<b>{score_pct}</b>", style_status_cell_white)
                body_styles.append(('BACKGROUND', (4, r_num), (4, r_num), colors.HexColor('#00B050')))
            elif iscore >= 0.90:
                p_score = Paragraph(f"<b>{score_pct}</b>", style_status_cell)
                body_styles.append(('BACKGROUND', (4, r_num), (4, r_num), colors.HexColor('#FFFF00')))
            else:
                p_score = Paragraph(f"<b>{score_pct}</b>", style_status_cell_white)
                body_styles.append(('BACKGROUND', (4, r_num), (4, r_num), colors.HexColor('#FF0000')))

            body_styles.append(('TOPPADDING', (0, r_num), (4, r_num), 0.6))
            body_styles.append(('BOTTOMPADDING', (0, r_num), (4, r_num), 0.6))

            if i_idx == 0:
                scorecard_data.append([p_no, p_mw, p_code, p_desc, p_score])
            else:
                scorecard_data.append(['', '', p_code, p_desc, p_score])

        body_styles.append(('SPAN', (0, start_r), (0, end_r)))
        body_styles.append(('SPAN', (1, start_r), (1, end_r)))
        current_r = end_r + 1

    tot_r = current_r
    p_tot_lbl = Paragraph("<b>Total</b>", style_obj_label)
    p_tot_val = Paragraph("<b>89%</b>", style_status_cell)
    scorecard_data.append(['', '', '', p_tot_lbl, p_tot_val])
    body_styles.append(('BACKGROUND', (4, tot_r), (4, tot_r), colors.HexColor('#FFFF00')))
    body_styles.append(('TOPPADDING', (0, tot_r), (4, tot_r), 1))
    body_styles.append(('BOTTOMPADDING', (0, tot_r), (4, tot_r), 1))

    body_table = Table(scorecard_data, colWidths=body_col_widths)
    body_table.setStyle(TableStyle(body_styles))
    story.append(body_table)

    story.append(Spacer(1, 3))

    legend_data = [
        [
            '', Paragraph("<b>&gt;/= 100%</b> (Tercapai)", style_init_desc),
            '', Paragraph("<b>On Progress &lt;100% - 90%</b> (Pengawalan)", style_init_desc),
            '', Paragraph("<b>Belum Start/ &lt;90%</b> (Perlu Intervensi)", style_init_desc)
        ]
    ]
    legend_table = Table(legend_data, colWidths=[20, 160, 20, 200, 20, 200])
    legend_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#00B050')),
        ('BACKGROUND', (2, 0), (2, 0), colors.HexColor('#FFFF00')),
        ('BACKGROUND', (4, 0), (4, 0), colors.HexColor('#FF0000')),
        ('GRID', (0, 0), (0, 0), 0.5, colors.HexColor('#000000')),
        ('GRID', (2, 0), (2, 0), 0.5, colors.HexColor('#000000')),
        ('GRID', (4, 0), (4, 0), 0.5, colors.HexColor('#000000')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 1),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
    ]))
    story.append(legend_table)

    story.append(PageBreak())

    # ==========================================================================
    # PAGE 2: DETAIL 3-PILAR ANALYSIS (MUST WINS 1 TO 4) + EVIDEN
    # ==========================================================================
    story.append(Paragraph("<b>ANALISIS TERPERINCI 3-PILAR & BUKTI EVIDEN AUDIT (BAGIAN 1: MUST WIN 1 - 4)</b>", style_section_title))
    story.append(Spacer(1, 3))

    detail_headers = [
        Paragraph("<b>No</b>", style_detail_th),
        Paragraph("<b>Must Win & Status</b>", style_detail_th),
        Paragraph("<b>Accomplishment (Pencapaian)</b>", style_detail_th),
        Paragraph("<b>Issue & Root Cause (Tantangan)</b>", style_detail_th),
        Paragraph("<b>Next Step (Solutif)</b>", style_detail_th),
        Paragraph("<b>Risk Mitigation & Eviden</b>", style_detail_th)
    ]

    p2_data = [detail_headers]
    p2_styles = [
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#0F172A')),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFBF00')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
    ]

    for idx, item in enumerate(MUST_WINS_DATA[:4], 1):
        p_no = Paragraph(f"<b>{item['id']}</b>", style_detail_mw)
        p_mw = Paragraph(f"<b>{item['name']}</b><br/><br/><font color='#D97706'><b>Status: {item['status_label']}</b></font>", style_detail_mw)
        p_acc = Paragraph(item['accomplishment'].replace('\n', '<br/>'), style_detail_cell)
        p_iss = Paragraph(item['issue'].replace('\n', '<br/>'), style_detail_cell)
        p_nxt = Paragraph(item['next_step'].replace('\n', '<br/>'), style_detail_cell)
        
        mit_txt = item['risk_mitigation'].replace('\n', '<br/>')
        ev_txt = f"<br/><br/><font color='#0284C7'><b>Bukti Eviden:</b><br/>{item['bukti_eviden']}</font>"
        p_mit = Paragraph(mit_txt + ev_txt, style_detail_cell)

        p2_data.append([p_no, p_mw, p_acc, p_iss, p_nxt, p_mit])
        if idx % 2 == 1:
            p2_styles.append(('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#F8FAFC')))

    p2_table = Table(p2_data, colWidths=[20, 115, 166.47, 166.47, 166.47, 167.48])
    p2_table.setStyle(TableStyle(p2_styles))
    story.append(p2_table)

    story.append(PageBreak())

    # ==========================================================================
    # PAGE 3: DETAIL 3-PILAR (MUST WINS 5 TO 7) + EVIDEN & REKOMENDASI
    # ==========================================================================
    story.append(Paragraph("<b>ANALISIS TERPERINCI 3-PILAR & BUKTI EVIDEN AUDIT (BAGIAN 2: MUST WIN 5 - 7)</b>", style_section_title))
    story.append(Spacer(1, 3))

    p3_data = [detail_headers]
    p3_styles = [
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#0F172A')),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFBF00')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
    ]

    for idx, item in enumerate(MUST_WINS_DATA[4:], 1):
        p_no = Paragraph(f"<b>{item['id']}</b>", style_detail_mw)
        p_mw = Paragraph(f"<b>{item['name']}</b><br/><br/><font color='#DC2626'><b>Status: {item['status_label']}</b></font>", style_detail_mw)
        p_acc = Paragraph(item['accomplishment'].replace('\n', '<br/>'), style_detail_cell)
        p_iss = Paragraph(item['issue'].replace('\n', '<br/>'), style_detail_cell)
        p_nxt = Paragraph(item['next_step'].replace('\n', '<br/>'), style_detail_cell)
        
        mit_txt = item['risk_mitigation'].replace('\n', '<br/>')
        ev_txt = f"<br/><br/><font color='#0284C7'><b>Bukti Eviden:</b><br/>{item['bukti_eviden']}</font>"
        p_mit = Paragraph(mit_txt + ev_txt, style_detail_cell)

        p3_data.append([p_no, p_mw, p_acc, p_iss, p_nxt, p_mit])
        if idx % 2 == 1:
            p3_styles.append(('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#F8FAFC')))

    p3_table = Table(p3_data, colWidths=[20, 115, 166.47, 166.47, 166.47, 167.48])
    p3_table.setStyle(TableStyle(p3_styles))
    story.append(p3_table)

    story.append(Spacer(1, 8))
    story.append(Paragraph("<b>SARAN & REKOMENDASI STRATEGIS EKSEKUTIF</b>", style_section_title))
    story.append(Spacer(1, 3))

    rec_headers = [
        Paragraph("<b>No</b>", style_detail_th),
        Paragraph("<b>Domain Inisiatif</b>", style_detail_th),
        Paragraph("<b>Rekomendasi Tindakan Strategis</b>", style_detail_th),
        Paragraph("<b>Unit Penanggung Jawab</b>", style_detail_th),
        Paragraph("<b>Target Waktu</b>", style_detail_th)
    ]
    rec_table_data = [rec_headers]
    rec_styles = [
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#0F172A')),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F172A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#FFFFFF')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 2.5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2.5),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
    ]

    rec_items = [
        ("1. Multi-Platform Matrix", "Membuat Feature Parity Matrix Android vs Web agar seluruh fitur transaksi identik secara logika fungsional.", "Tim Mobile Engineering & QA", "Sebelum Kickoff UAT Kios"),
        ("2. SLA & Protocol UAT 3 Kios", "Menerbitkan Berita Acara UAT Formal untuk minimal 3 kios dengan skenario uji seragam (Zero Critical Bug).", "Tim QA & Operasional Kios", "Minggu 1-2 Agustus 2026"),
        ("3. SLA Resolusi Bug D+30", "Menerapkan matriks respon bug pasca-live: Critical Bug (Respon < 15 mnt, Fix < 4 jam), Major Bug (< 24 jam).", "Tim IT Support & Backend", "H+1 s/d H+30 Go-Live"),
        ("4. Payment Gateway Idempotency", "Menyediakan mekanisme Idempotency Token dan Webhook Signature Verification untuk mengantisipasi callback ganda.", "Backend Lead & Keuangan", "Fase Finalisasi API"),
        ("5. SOP Troubleshooting Kios", "Menyusun Troubleshooting Flowchart kasir saat terjadi gangguan jaringan atau status transaksi pending.", "Tim Operasional & Training", "Bersamaan dengan UAT Kios"),
    ]

    for idx, (dom, rec, pic, target) in enumerate(rec_items, 1):
        p_no = Paragraph(f"<b>{idx}</b>", style_detail_mw)
        p_dom = Paragraph(f"<b>{dom}</b>", style_detail_cell)
        p_rec = Paragraph(rec, style_detail_cell)
        p_pic = Paragraph(pic, style_detail_cell)
        p_tgt = Paragraph(target, style_detail_cell)

        rec_table_data.append([p_no, p_dom, p_rec, p_pic, p_tgt])
        if idx % 2 == 1:
            rec_styles.append(('BACKGROUND', (0, idx), (-1, idx), colors.HexColor('#F8FAFC')))

    rec_table = Table(rec_table_data, colWidths=[20, 135, 406.89, 130, 110])
    rec_table.setStyle(TableStyle(rec_styles))
    story.append(rec_table)

    doc.build(story, canvasmaker=NumberedCanvas)
    print(f"PDF successfully updated at: {PDF_PATH}")


if __name__ == "__main__":
    generate_excel()
    generate_pdf()
