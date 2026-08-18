import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
wb.remove(wb.active)

# Styling constants
font_title = Font(name='Segoe UI', size=13, bold=True, color='FFFFFF')
font_header = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
font_bold = Font(name='Segoe UI', size=9.5, bold=True, color='000000')
font_regular = Font(name='Segoe UI', size=9.5, color='000000')
font_muted = Font(name='Segoe UI', size=8.5, italic=True, color='555555')

fill_navy_header = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
fill_orange_header = PatternFill(start_color='EA580C', end_color='EA580C', fill_type='solid')
fill_meta = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
fill_green = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
fill_orange = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
fill_red = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')

thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

align_center = Alignment(horizontal='center', vertical='center')
align_left = Alignment(horizontal='left', vertical='center')
align_right = Alignment(horizontal='right', vertical='center')

# ==================== SHEET 1: MASTER DASHBOARD ====================
ws1 = wb.create_sheet(title="BSC_Master_Dashboard")
ws1.views.sheetView[0].showGridLines = True

# Title Header
ws1.merge_cells('A1:F1')
ws1['A1'] = "DASHBOARD EKSEKUTIF: 9 BALANCED SCORECARDS (BSC) KARISMA ONLINE"
ws1['A1'].font = font_title
ws1['A1'].fill = fill_navy_header
ws1['A1'].alignment = align_center
ws1.row_dimensions[1].height = 28

ws1.merge_cells('A2:F2')
ws1['A2'] = "Evaluasi Kinerja Terpadu Berbasis Sumber Data Tunggal (Single Source of Truth)"
ws1['A2'].font = Font(name='Segoe UI', size=9.5, italic=True, color='FFFFFF')
ws1['A2'].fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
ws1['A2'].alignment = align_center
ws1.row_dimensions[2].height = 18

# Meta Bar
ws1['A4'] = "Periode Evaluasi:"
ws1['B4'] = "Q3 2026 (Launch Readiness)"
ws1['D4'] = "Rata-Rata 9 BSC:"
ws1['E4'] = "=AVERAGE(E7:E15)"

for col in ['A4', 'B4', 'D4', 'E4']:
    ws1[col].font = font_bold
    ws1[col].fill = fill_meta
    ws1[col].border = thin_border
ws1['E4'].font = Font(name='Segoe UI', size=11, bold=True, color='047857')
ws1['E4'].number_format = '0.0%'
ws1['E4'].alignment = align_center

# Master Table
headers1 = ["No", "Perspektif BSC", "Unit Penanggung Jawab (Owner)", "Measure Lead Utama", "Overall Status", "Indikator Kinerja"]
for col_idx, h in enumerate(headers1, 1):
    cell = ws1.cell(row=6, column=col_idx)
    cell.value = h
    cell.font = font_header
    cell.fill = fill_orange_header
    cell.alignment = align_center
    cell.border = thin_border
ws1.row_dimensions[6].height = 22

dashboard_rows = [
    (1, "1. BSC Karisma Online", "Tim Digital & IT Karisma Online", "Status Kesiapan API, Test Coverage, Kepatuhan App Store", 0.95, "■ On Progress (90-99%)"),
    (2, "2. BSC Sisi Customer", "Tim Customer Experience & Komersial", "Indeks Rating Layanan Sales, Retensi Mitra, Siklus Order", 0.94, "■ On Progress (90-99%)"),
    (3, "3. BSC Divisi Sales", "Kepala Divisi Penjualan & Distribusi", "Utilisasi Limit Kredit Toko, Rasio Toko Aktif Order", 0.89, "■ Perlu Intervensi (<90%)"),
    (4, "4. BSC Logistik & Warehouse", "Tim Logistik & Operasional Gudang", "SLA Pengemasan (<2 Jam), Akurasi Ongkir Kecamatan", 0.92, "■ On Progress (90-99%)"),
    (5, "5. BSC Sisi Multiplatform", "Tim Mobile Engineering & UI/UX", "App Store Review Clearance, Android Crash Rate (<0.1%)", 0.96, "■ On Progress (90-99%)"),
    (6, "6. BSC Sisi Perusahaan", "Dewan Direksi & Tim Manajemen Eksekutif", "Days Sales Outstanding (DSO), Efisiensi Biaya per Order", 0.94, "■ On Progress (90-99%)"),
    (7, "7. BSC Payment & Keuangan", "Divisi Keuangan & Treasury Karisma", "Adopsi BRIVA, Kecepatan Rekonsiliasi Kas, Rasio Over-Limit", 0.96, "■ On Progress (90-99%)"),
    (8, "8. BSC Sisi Teknis", "Tim Lead Backend Engineering", "API Latency (<200ms), Error Rate (<0.01%), ACID Transaksi", 0.95, "■ On Progress (90-99%)"),
    (9, "9. BSC Sisi Keamanan", "Tim Security & Compliance", "Zero Critical Vulnerability, Token Expiry, Audit Non-PII", 0.97, "■ On Progress (90-99%)")
]

for idx, (no, pers, owner, lead, score, st) in enumerate(dashboard_rows, 7):
    ws1.cell(row=idx, column=1, value=no).alignment = align_center
    ws1.cell(row=idx, column=2, value=pers).alignment = align_left
    ws1.cell(row=idx, column=3, value=owner).alignment = align_left
    ws1.cell(row=idx, column=4, value=lead).alignment = align_left
    
    c5 = ws1.cell(row=idx, column=5, value=score)
    c5.alignment = align_center
    c5.number_format = '0%'
    c5.font = font_bold
    
    c6 = ws1.cell(row=idx, column=6, value=st)
    c6.alignment = align_center
    c6.font = font_bold
    
    for c in range(1, 7):
        ws1.cell(row=idx, column=c).border = thin_border
        
    if score >= 0.90:
        c5.fill = fill_orange
        c6.fill = fill_orange
    else:
        c5.fill = fill_red
        c6.fill = fill_red
    ws1.row_dimensions[idx].height = 20

# Integrity Section
ws1.merge_cells('A18:F18')
ws1['A18'] = "PRINSIP INTEGRITAS DATA & SUMBER DOKUMENTASI RESMI:"
ws1['A18'].font = font_bold
ws1['A18'].fill = fill_meta

int_rows = [
    ("1. Arsitektur & Kepatuhan Platform (95%)", "Terverifikasi pada controller mobile CodeIgniter (Mobile.php) dan arsitektur iOS Swift/Android Kotlin. Lolos Apple Review 5.1.1 melalui audit tabel mobile_account_deletions."),
    ("2. Finansial & Limit Kredit Pelanggan (94%)", "Data tagihan berjalan dan limit kredit dibaca langsung dari customers.max_credit dan agregasi tabel orders serta view database v_products."),
    ("3. Kesiapan Divisi Sales & Distribusi (89%)", "Pemetaan akun toko binaan aktif di backend (customers.salesman_id). Tantangan utama terletak pada legalitas skema komisi teritori digital untuk mendorong adopsi lapangan."),
    ("4. Ekosistem Pembayaran & Virtual Account (96%)", "Kanal BRIVA aktif via library Brivaws pada tabel briva_api dengan auto-update status pembayaran instan ke pesanan lunas (status = 10).")
]

for idx, (t, d) in enumerate(int_rows, 19):
    ws1.cell(row=idx, column=1, value=t).font = font_bold
    ws1.cell(row=idx, column=1).fill = fill_meta
    ws1.cell(row=idx, column=1).border = thin_border
    
    ws1.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=6)
    ws1.cell(row=idx, column=2, value=d).font = font_regular
    ws1.cell(row=idx, column=2).alignment = align_left
    for c in range(2, 7):
        ws1.cell(row=idx, column=c).border = thin_border
    ws1.row_dimensions[idx].height = 20

ws1.column_dimensions['A'].width = 6
ws1.column_dimensions['B'].width = 28
ws1.column_dimensions['C'].width = 32
ws1.column_dimensions['D'].width = 42
ws1.column_dimensions['E'].width = 14
ws1.column_dimensions['F'].width = 24

# ==================== HELPER FOR 9 PERSPECTIVE SHEETS WITH MERGED MUST WIN ====================
def create_perspective_sheet(sheet_title, bsc_name, obj_desc, meta_dict, initiatives, total_score):
    ws = wb.create_sheet(title=sheet_title)
    ws.views.sheetView[0].showGridLines = True
    
    # Header Banner
    ws.merge_cells('A1:D1')
    ws['A1'] = f"BALANCED SCORECARD: {bsc_name.upper()}"
    ws['A1'].font = font_title
    ws['A1'].fill = fill_navy_header
    ws['A1'].alignment = align_center
    ws.row_dimensions[1].height = 24
    
    # Objective Description
    ws['A2'] = "Objective Description"
    ws['A2'].font = font_header
    ws['A2'].fill = fill_navy_header
    ws['A2'].alignment = align_center
    
    ws.merge_cells('B2:D2')
    ws['B2'] = obj_desc
    ws['B2'].font = font_regular
    ws['B2'].alignment = align_left
    for c in range(1, 5):
        ws.cell(row=2, column=c).border = thin_border
    ws.row_dimensions[2].height = 28
    
    # Meta Bar
    ws['A3'] = f"Owner: {meta_dict['owner']}"
    ws['B3'] = f"Measure lead: {meta_dict['lead']}"
    ws['C3'] = f"Frequency: {meta_dict['freq']} | Overall Status: {meta_dict['status']}"
    ws['D3'] = meta_dict['date']
    
    for c in range(1, 5):
        ws.cell(row=3, column=c).font = font_bold
        ws.cell(row=3, column=c).fill = fill_meta
        ws.cell(row=3, column=c).border = thin_border
    ws['D3'].fill = fill_navy_header
    ws['D3'].font = font_header
    ws['D3'].alignment = align_center
    ws.row_dimensions[3].height = 20
    
    # Table Header
    headers = ["Must Win", "Key initiatives", "Dasar Verifikasi / Sumber Data", "Status"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = h
        cell.font = font_header
        cell.fill = fill_orange_header
        cell.alignment = align_center
        cell.border = thin_border
    ws.row_dimensions[5].height = 22
    
    # Group consecutive initiatives with the same Must Win
    groups = []
    current_mw = None
    current_start = 6
    
    start_row = 6
    for idx, (mw, init, audit, st_val) in enumerate(initiatives, start_row):
        ws.cell(row=idx, column=1, value=mw).font = font_bold
        ws.cell(row=idx, column=1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        ws.cell(row=idx, column=2, value=init).font = font_regular
        ws.cell(row=idx, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        ws.cell(row=idx, column=3, value=audit).font = font_muted
        ws.cell(row=idx, column=3).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        c4 = ws.cell(row=idx, column=4, value=st_val)
        c4.font = font_bold
        c4.alignment = align_center
        c4.number_format = '0%'
        
        if st_val >= 1.0:
            c4.fill = fill_green
        elif st_val >= 0.90:
            c4.fill = fill_orange
        else:
            c4.fill = fill_red
            
        for c in range(1, 5):
            ws.cell(row=idx, column=c).border = thin_border
        ws.row_dimensions[idx].height = 22
        
        if current_mw is None:
            current_mw = mw
            current_start = idx
        elif mw != current_mw:
            groups.append((current_start, idx - 1))
            current_mw = mw
            current_start = idx
            
    if current_mw is not None:
        groups.append((current_start, start_row + len(initiatives) - 1))
        
    # MERGE CELLS FOR MUST WIN COLUMN (COLUMN A)
    for g_start, g_end in groups:
        if g_end > g_start:
            ws.merge_cells(start_row=g_start, start_column=1, end_row=g_end, end_column=1)
            for r in range(g_start, g_end + 1):
                ws.cell(row=r, column=1).border = thin_border
            ws.cell(row=g_start, column=1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws.cell(row=g_start, column=1).font = font_bold
        
    tot_row = start_row + len(initiatives)
    ws.cell(row=tot_row, column=1, value="").fill = fill_navy_header
    ws.cell(row=tot_row, column=2, value="").fill = fill_navy_header
    ws.cell(row=tot_row, column=3, value="Total Capaian Kinerja Terverifikasi (Overall Score)").font = font_header
    ws.cell(row=tot_row, column=3).fill = fill_navy_header
    ws.cell(row=tot_row, column=3).alignment = align_right
    
    c_tot = ws.cell(row=tot_row, column=4, value=total_score)
    c_tot.font = font_header
    c_tot.alignment = align_center
    c_tot.fill = fill_orange_header if total_score >= 0.90 else fill_red
    c_tot.number_format = '0%'
    
    for c in range(1, 5):
        ws.cell(row=tot_row, column=c).border = thin_border
    ws.row_dimensions[tot_row].height = 22
    
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 52
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 14

# 1. Apps
create_perspective_sheet(
    "1_Karisma_Online",
    "Karisma Online (Platform & Digital Ecosystem)",
    "Digitalisasi menyeluruh kanal pemesanan B2B/B2C, integrasi monitoring limit kredit & penagihan virtual account (BRIVA), percepatan pemenuhan order, serta kesiapan distribusi multi-platform yang aman dan patuh regulasi.",
    {"owner": "Tim Digital & IT Karisma Online", "lead": "Status Kesiapan API, Test Coverage, Kepatuhan App Store", "freq": "Bulanan", "status": "95%", "date": "31-Agu-26"},
    [
        ("1. KEANDALAN AKSES & PENGALAMAN PENGGUNA", "1.1 Akses pemesanan mandiri 24/7 (Aplikasi iOS Swift, Android & Web KiuStore)", "Terverifikasi: MainTabView, Catalog, Cart, Checkout", 1.00),
        ("1. KEANDALAN AKSES & PENGALAMAN PENGGUNA", "1.2 Mode jelajah katalog produk & promo tanpa kewajiban login awal (Guest Browsing)", "Terverifikasi: GuestAccessTests & LoginView", 1.00),
        ("1. KEANDALAN AKSES & PENGALAMAN PENGGUNA", "1.3 Sinkronisasi harga grosir & stok produk real-time via API /api/v1", "Terverifikasi: APIEndpoint & Staging API", 0.95),
        ("2. INTEGRASI FINANSIAL & TRANSAKSI", "2.1 Integrasi pembayaran Virtual Account instan BRIVA via Brivaws", "Terverifikasi: MobileBrivaPaymentAPI & OrderTests", 0.95),
        ("2. INTEGRASI FINANSIAL & TRANSAKSI", "2.2 Visibilitas limit kredit (max_credit) & tagihan berjalan real-time bagi mitra", "Terverifikasi: CustomerFinanceTests & GET /customer/finance-summary", 1.00),
        ("2. INTEGRASI FINANSIAL & TRANSAKSI", "2.3 Otomasi validasi persetujuan pesanan sesuai plafon kredit toko (payment_method = 1)", "Terverifikasi: CheckoutViewModelTests", 1.00),
        ("3. KECEPATAN LOGISTIK & ORDER", "3.1 Pemrosesan alur pesanan dari checkout hingga pengiriman (Order SLA)", "Terverifikasi: Alur status pesanan 1, 2, 3, 4, 5", 0.90),
        ("3. KECEPATAN LOGISTIK & ORDER", "3.2 Transparansi status kirim, ongkir (mobile_shipping_quotes) & faktur digital", "Terverifikasi: ShippingTests & InvoiceView", 0.95),
        ("3. KECEPATAN LOGISTIK & ORDER", "3.3 Fitur pemesanan cepat rutin 1-sentuhan (Quick Re-Order)", "Status: Dalam pengembangan modul", 0.85),
        ("4. KEPATUHAN REGULASI & DATA", "4.1 Kepatuhan 100% standar Apple App Review & Privasi (Hapus Akun DELETE /account)", "Terverifikasi: AppReviewResolution & ProductionSafetyTests", 1.00),
        ("4. KEPATUHAN REGULASI & DATA", "4.2 Layanan bantuan pelanggan interaktif terintegrasi (In-App Chat Support)", "Terverifikasi: ChatTests & ChatView", 0.95),
        ("4. KEPATUHAN REGULASI & DATA", "4.3 Pembentukan duta digital (Digital Champion) di cabang untuk pendampingan toko", "Status: Menunggu standardisasi SOP cabang", 0.80)
    ],
    0.95
)

# 2. Customer
create_perspective_sheet(
    "2_Customer",
    "Sisi Customer (Customer Experience & Mitra Access)",
    "Memberikan kemudahan pemesanan mandiri 24/7, fleksibilitas tingkatan harga sesuai level mitra, transparansi kupon diskon, layanan interaktif responsif, dan keterbukaan ulasan kepuasan produk.",
    {"owner": "Tim Customer Experience & Komersial", "lead": "Indeks Rating Layanan Sales, Retensi Mitra, Siklus Order", "freq": "Bulanan", "status": "94%", "date": "31-Agu-26"},
    [
        ("1. KEMUDAHAN AKSES & BELANJA", "1.1 Mode Guest Browsing katalog produk tanpa kewajiban login awal", "Terverifikasi: Endpoint GET /api/v1/products & UI Katalog", 1.00),
        ("1. KEMUDAHAN AKSES & BELANJA", "1.2 Manajemen data profil toko, kontak, dan multi-alamat pengiriman", "Terverifikasi: Endpoint PUT /api/v1/profile & Tabel customers", 1.00),
        ("1. KEMUDAHAN AKSES & BELANJA", "1.3 Sistem kupon promo potongan harga transaksi belanja", "Terverifikasi: Tabel coupons & validasi Shop.php", 0.95),
        ("2. LAYANAN INTERAKTIF & CHAT", "2.1 Saluran pesan langsung Customer Service 2 arah (In-App Live Chat)", "Terverifikasi: Endpoint /api/v1/messages & Tabel message", 0.95),
        ("2. LAYANAN INTERAKTIF & CHAT", "2.2 Sistem ulasan dan rating bintang kualitas produk pasca pesanan selesai", "Terverifikasi: Endpoint POST /api/v1/orders/{id}/complete", 0.95),
        ("2. LAYANAN INTERAKTIF & CHAT", "2.3 Notifikasi pengingat pembayaran & status pesanan instan", "Status: Integrasi gateway notifikasi pesan", 0.85),
        ("3. TRANSPARANSI & LOYALITAS", "3.1 Akses riwayat faktur, rincian biaya, dan status pengiriman real-time", "Terverifikasi: Endpoint GET /api/v1/orders & Tabel orders", 1.00),
        ("3. TRANSPARANSI & LOYALITAS", "3.2 Program poin loyalitas dan reward kuota diskon toko aktif", "Status: Tahap perancangan skema reward Q4 2026", 0.80)
    ],
    0.94
)

# 3. Sales
create_perspective_sheet(
    "3_Sales",
    "Sisi Sales (Sales Force Enablement & Distribusi)",
    "Memberdayakan tenaga penjual lapangan, mengoptimalkan pembagian wilayah kios binaan, mengawal kepatuhan limit piutang, dan mempercepat rekonsiliasi order digital.",
    {"owner": "Kepala Divisi Penjualan & Distribusi", "lead": "Utilisasi Limit Kredit Toko, Rasio Toko Aktif Order", "freq": "Bulanan", "status": "89%", "date": "31-Agu-26"},
    [
        ("1. PEMETAAN WILAYAH & BINAAN", "1.1 Penugasan kios binaan per salesman lapangan di basis data terpusat", "Terverifikasi: Field customers.salesman_id & Admin Salesman", 1.00),
        ("1. PEMETAAN WILAYAH & BINAAN", "1.2 Visibilitas riwayat transaksi dan status pembayaran toko binaan", "Terverifikasi: Modul Admin Salesman API & Controller Orders", 0.95),
        ("1. PEMETAAN WILAYAH & BINAAN", "1.3 Perencanaan rute kunjungan sales terintegrasi riwayat order kios", "Status: Dalam perancangan modul territory route", 0.80),
        ("2. PENGAWALAN PLAFON KREDIT", "2.1 Visibilitas limit kredit (max_credit) toko binaan saat pemesanan", "Terverifikasi: Model Mobile_api_model & Admin Piutang", 0.95),
        ("2. PENGAWALAN PLAFON KREDIT", "2.2 Alur validasi persetujuan pesanan tempo digital oleh supervisor sales", "Terverifikasi: Status alur piutang di Piutang.php", 0.90),
        ("2. PENGAWALAN PLAFON KREDIT", "2.3 Skema insentif komisi teritori digital untuk mendorong adopsi mobile", "Status: Menunggu legalitas formal skema komisi direksi", 0.80),
        ("3. PENDAMPINGAN & KUALITAS", "3.1 Program pendampingan instalasi aplikasi ke kios mitra perintis", "Terverifikasi: Pelaksanaan onboarding tim cabang", 0.85),
        ("3. PENDAMPINGAN & KUALITAS", "3.2 Sistem penilaian kinerja pelayanan salesman dari ulasan toko", "Terverifikasi: Modul Rating.php & Tabel reviews", 0.95)
    ],
    0.89
)

# 4. Logistik
create_perspective_sheet(
    "4_Logistik_Warehouse",
    "Sisi Logistik & Warehouse (Supply Chain & SLA)",
    "Menjamin presisi perhitungan ongkos kirim hingga tingkat kecamatan, efisiensi alur pengemasan gudang (Order SLA), mitigasi produk lambat bergerak (deadstock), dan penyediaan armada muatan besar.",
    {"owner": "Tim Logistik & Operasional Gudang", "lead": "SLA Pengemasan (<2 Jam), Akurasi Ongkir, Rasio Deadstock", "freq": "Harian", "status": "92%", "date": "31-Agu-26"},
    [
        ("1. PRESISI LOGISTIK WILAYAH", "1.1 Integrasi pembacaan tarif ekspedisi tingkat kecamatan (RajaOngkir Pro API)", "Terverifikasi: Endpoint shipping/destination & Rajaongkir.php", 1.00),
        ("1. PRESISI LOGISTIK WILAYAH", "1.2 Penguncian kuotasi tarif ongkos kirim 30 menit saat proses checkout", "Terverifikasi: Tabel mobile_shipping_quotes & expiry lock", 1.00),
        ("1. PRESISI LOGISTIK WILAYAH", "1.3 Kalkulasi otomatis akumulasi berat produk satuan botol/pcs dan dus/karton", "Terverifikasi: Model validasi berat & product_unit_value", 1.00),
        ("2. KECEPATAN PEMENUHAN GUDANG", "2.1 Antrean status alur pesanan terstruktur (Verifikasi -> Kemas -> Kirim)", "Terverifikasi: Controller Pengiriman.php & Admin Orders", 0.95),
        ("2. KECEPATAN PEMENUHAN GUDANG", "2.2 Standarisasi target penyiapan barang gudang seragam (SLA < 2 Jam)", "Status: Penyelarasan SOP shift tim gudang cabang", 0.85),
        ("2. KECEPATAN PEMENUHAN GUDANG", "2.3 Skema pengiriman armada truk internal Karisma untuk muatan tonase besar", "Status: Penentuan tarif flat rute rutin armada internal", 0.80),
        ("3. PENGENDALIAN DEADSTOCK", "3.1 Deteksi dini produk slow-moving (> 1 tahun pergerakan < 5% & retur)", "Terverifikasi: Catatan teknis CATATAN PENTING.txt #DEADSTOCK", 0.90),
        ("3. PENGENDALIAN DEADSTOCK", "3.2 Otomasi alur diskon kilat (flash sale) cuci gudang produk deadstock", "Status: Tahap sinkronisasi formula promo staging", 0.80)
    ],
    0.92
)

# 5. Multiplatform
create_perspective_sheet(
    "5_Multiplatform",
    "Sisi Multiplatform (Tri-Platform iOS, Android, Web)",
    "Menghadirkan performa aplikasi native berkecepatan tinggi, ukuran aplikasi hemat kuota, antarmuka responsif di ponsel/tablet, dan stabilitas back-office web portal.",
    {"owner": "Tim Mobile Engineering & UI/UX", "lead": "App Store Review Clearance, Android Crash Rate (<0.1%)", "freq": "Bulanan", "status": "96%", "date": "31-Agu-26"},
    [
        ("1. EKOSISTEM IOS (APP STORE)", "1.1 Arsitektur Swift Native (SwiftUI & UIKit) berkinerja tinggi", "Terverifikasi: Source build iOS Karisma Online", 1.00),
        ("1. EKOSISTEM IOS (APP STORE)", "1.2 Kepatuhan Apple Review Guideline 5.1.1(v) (Penghapusan Akun Mandiri)", "Terverifikasi: Dokumen docs/APP_REVIEW_RESOLUTION_20260728.md", 1.00),
        ("1. EKOSISTEM IOS (APP STORE)", "1.3 Tata letak antarmuka adaptif layar iPad dengan tombol navigasi jelas", "Terverifikasi: Kepatuhan review UI iPad resolution", 0.95),
        ("2. EKOSISTEM ANDROID (PLAY STORE)", "2.1 Arsitektur Kotlin/Java Native (MVVM Pattern) responsif", "Terverifikasi: Source code kiustore_apps module", 1.00),
        ("2. EKOSISTEM ANDROID (PLAY STORE)", "2.2 Optimasi ukuran installer APK hemat kuota internet (< 15MB)", "Terverifikasi: Build packaging & resource shrinking", 0.95),
        ("2. EKOSISTEM ANDROID (PLAY STORE)", "2.3 Kompresi otomatis foto bukti transfer bank sebelum diunggah", "Terverifikasi: Controller Mobile.php payment_picture_name", 0.95),
        ("3. PORTAL BACK-OFFICE WEB", "3.1 Dasbor operasional web AdminLTE dengan manajemen modul lengkap", "Terverifikasi: Module application/modules/admin", 1.00),
        ("3. PORTAL BACK-OFFICE WEB", "3.2 Sinkronisasi status transaksi lintas platform (iOS, Android, Web)", "Terverifikasi: Database MariaDB InnoDB ACID transactions", 0.95)
    ],
    0.96
)

# 6. Perusahaan
create_perspective_sheet(
    "6_Perusahaan",
    "Sisi Perusahaan (Governance & Corporate Finance)",
    "Mengakselerasi perputaran modal kerja, menekan Days Sales Outstanding (DSO), menjamin kepatuhan audit perpajakan, dan meningkatkan efisiensi biaya operasional per order.",
    {"owner": "Dewan Direksi & Tim Manajemen Eksekutif", "lead": "Days Sales Outstanding (DSO), Efisiensi Biaya per Order", "freq": "Bulanan", "status": "94%", "date": "31-Agu-26"},
    [
        ("1. ARUS KAS & MODAL KERJA", "1.1 Pemotongan waktu settlement pembayaran transfer dari 2-4 jam ke < 5 detik", "Terverifikasi: Otomasi webhook callback BRIVA", 1.00),
        ("1. ARUS KAS & MODAL KERJA", "1.2 Penurunan risiko piutang tak tertagih via validasi limit plafon kredit sistem", "Terverifikasi: Logika max_credit di checkout mobile", 0.95),
        ("1. ARUS KAS & MODAL KERJA", "1.3 Visibilitas dasbor eksekutif pergerakan omset harian dan bulanan", "Terverifikasi: Controller Dashboard.php & Report.php", 0.95),
        ("2. TATA KELOLA HUKUM & AUDIT", "2.1 Pemisahan data identitas pribadi (non-PII) tanpa menghapus faktur transaksi", "Terverifikasi: Model Mobile_api_model.php delete_account", 1.00),
        ("2. TATA KELOLA HUKUM & AUDIT", "2.2 Perlindungan integritas nomor faktur dan pesanan untuk audit perpajakan", "Terverifikasi: Relasi tabel orders, order_items, payments", 1.00),
        ("2. TATA KELOLA HUKUM & AUDIT", "2.3 Pencatatan audit trail lengkap seluruh aktivitas transaksi dan pembayaran", "Terverifikasi: Tabel briva_api, mobile_account_deletions", 0.95),
        ("3. SKALABILITAS KORPORASI", "3.1 Efisiensi biaya operasional per transaksi pemesanan", "Terverifikasi: Otomasi sistem mengurangi beban kasir", 0.90),
        ("3. SKALABILITAS KORPORASI", "3.2 Standardisasi SOP operasional cabang untuk ekspansi digital", "Status: Penyusunan panduan operasional cabang terpadu", 0.85)
    ],
    0.94
)

# 7. Keuangan
create_perspective_sheet(
    "7_Keuangan_Payment",
    "Sisi Keuangan & Payment (Fintech & Pricing Engine)",
    "Otomasi penerimaan pembayaran Virtual Account (BRIVA), perlindungan skema harga bertingkat, kepastian rekonsiliasi kas harian, dan pencegahan transaksi over-limit.",
    {"owner": "Divisi Keuangan & Treasury Karisma", "lead": "Adopsi BRIVA, Kecepatan Rekonsiliasi Kas, Rasio Over-Limit", "freq": "Harian", "status": "96%", "date": "31-Agu-26"},
    [
        ("1. OTOMASI FINTECH BRIVA", "1.1 Integrasi library Brivaws standar SNAP API BRI (/transfer-va/create-va)", "Terverifikasi: Controller Brivawsapi.php & key/private.pem", 0.95),
        ("1. OTOMASI FINTECH BRIVA", "1.2 Penerbitan Dynamic VA (91118 + No HP) dengan batas kedaluwarsa 15 menit", "Terverifikasi: CATATAN PENTING.txt (V) & Tabel briva_api", 1.00),
        ("1. OTOMASI FINTECH BRIVA", "1.3 Webhook callback otomatis memperbarui status pesanan menjadi Lunas (status = 10)", "Terverifikasi: Function inquiryVa & callback handler", 0.95),
        ("2. PROTEKSI HARGA MULTI-TIER", "2.1 Proteksi skema harga 3 level (Level 1: Retail, Level 2: Grosir, Level 3: Distributor)", "Terverifikasi: Database View v_products & level_product", 1.00),
        ("2. PROTEKSI HARGA MULTI-TIER", "2.2 Sistem satuan ganda (Botol/Pcs ke Dus/Karton dengan konversi otomatis)", "Terverifikasi: Modul keranjang unit_type 1 & 2", 1.00),
        ("2. PROTEKSI HARGA MULTI-TIER", "2.3 Kalkulasi potongan promo dan diskon kupon langsung di sisi server", "Terverifikasi: Query promo aktif CAST(expired_date AS DATE)", 0.95),
        ("3. KONTROL LIMIT KREDIT", "3.1 Validasi otomatis pencegahan pemesanan melebihi limit plafon (max_credit)", "Terverifikasi: Model Mobile_api_model checkout validation", 1.00),
        ("3. KONTROL LIMIT KREDIT", "3.2 Jalur pembayaran cadangan Transfer Manual Multi-Bank dengan verifikasi kasir", "Terverifikasi: Endpoint POST /api/v1/orders/{id}/confirm-transfer", 0.90)
    ],
    0.96
)

# 8. Teknis
create_perspective_sheet(
    "8_Teknis",
    "Sisi Teknis (Backend Architecture & RESTful API)",
    "Menyediakan infrastruktur backend RESTful API yang cepat (<200ms), modular (HMVC), aman, kompatibel PHP 8.x, dengan integritas transaksi MariaDB ACID.",
    {"owner": "Tim Lead Backend Engineering", "lead": "API Latency (<200ms), Error Rate (<0.01%), ACID Transaksi DB", "freq": "Bulanan", "status": "95%", "date": "31-Agu-26"},
    [
        ("1. RESTFUL API ARCHITECTURE", "1.1 Endpoint API /api/v1 terisolasi pada modul application/modules/api", "Terverifikasi: Controller Mobile.php (827 baris kode)", 1.00),
        ("1. RESTFUL API ARCHITECTURE", "1.2 Penanganan payload JSON murni dan standarisasi response error HTTP", "Terverifikasi: Function respond & error di Mobile.php", 1.00),
        ("1. RESTFUL API ARCHITECTURE", "1.3 Konfigurasi CORS header (Access-Control-Allow-Origin: *) & HTTP verbs", "Terverifikasi: Preflight OPTIONS handler di Mobile.php", 1.00),
        ("2. BASIS DATA & TRANSAKSI", "2.1 Migration scripts database (20260629_mobile_api, 20260728_account_deletion)", "Terverifikasi: Folder db/migrations/*.sql", 1.00),
        ("2. BASIS DATA & TRANSAKSI", "2.2 Penerapan transaksi ACID (trans_begin, trans_commit, trans_rollback)", "Terverifikasi: Model Mobile_api_model.php (1612 baris kode)", 1.00),
        ("2. BASIS DATA & TRANSAKSI", "2.3 Optimasi query view v_products dan pemanfaatan indeks tabel relasi", "Terverifikasi: sql_view.sql & indeks foreign key dump", 0.95),
        ("3. KINERJA & SKALABILITAS", "3.1 Kompatibilitas penuh CodeIgniter 3 HMVC pada lingkungan PHP 8.x", "Terverifikasi: Log error testing & library compatibility", 0.95),
        ("3. KINERJA & SKALABILITAS", "3.2 Keranjang belanja mobile berbasis basis data tanpa dependensi sesi web", "Terverifikasi: Tabel mobile_cart_items", 1.00),
        ("3. KINERJA & SKALABILITAS", "3.3 Cakupan pengujian unit test API otomatis", "Status: Perluasan skenario pengujian beban transaksi", 0.85)
    ],
    0.95
)

# 9. Keamanan
create_perspective_sheet(
    "9_Keamanan",
    "Sisi Keamanan (Security & Compliance)",
    "Menjamin keamanan data kredensial pengguna, enkripsi token sesi, perlindungan hak privasi non-PII, dan kepatuhan audit regulasi global.",
    {"owner": "Tim Security & Compliance", "lead": "Zero Critical Vulnerability, Token Expiry, Audit Trail Non-PII", "freq": "Bulanan", "status": "97%", "date": "31-Agu-26"},
    [
        ("1. AUTENTIKASI & SESI", "1.1 Bearer Token SHA-256 (mobile_api_tokens) dengan kedaluwarsa 30 hari", "Terverifikasi: issue_token & user_from_token SHA-256", 1.00),
        ("1. AUTENTIKASI & SESI", "1.2 Hashing kata sandi pengguna berstandar industri BCRYPT (password_hash)", "Terverifikasi: Register & login verify di Mobile.php", 1.00),
        ("1. AUTENTIKASI & SESI", "1.3 Pencabutan token instan saat logout dan pembatasan hak akses level", "Terverifikasi: Function revoke_token & require_fields", 1.00),
        ("2. PRIVASI & NON-PII", "2.1 Endpoint penghapusan akun mandiri DELETE /api/v1/account", "Terverifikasi: App Store Guideline 5.1.1(v) compliance", 1.00),
        ("2. PRIVASI & NON-PII", "2.2 Audit trail non-PII dengan hash SHA-256 pada tabel mobile_account_deletions", "Terverifikasi: Struktur kolom email_hash CHAR(64)", 1.00),
        ("2. PRIVASI & NON-PII", "2.3 Pelepasan relasi PII pada pesanan tanpa menghapus arsip pembukuan", "Terverifikasi: Anonymization logic di delete_account", 1.00),
        ("3. OPERASIONAL & AUDIT", "3.1 Isolasi akun uji coba internal (is_internal) agar tidak mengotori omset riil", "Terverifikasi: Pemisahan data akun demo di database", 0.95),
        ("3. OPERASIONAL & AUDIT", "3.2 Sanitasi input payload JSON mencegah ancaman SQL Injection & XSS", "Terverifikasi: Input casting & query bindings", 0.95),
        ("3. OPERASIONAL & AUDIT", "3.3 Pembaruan rutin sertifikat keamanan SSL/TLS dan private key perbankan", "Terverifikasi: File key/private.pem & HTTPS headers", 0.90)
    ],
    0.97
)

# ==================== SHEET 11: MATRIKS GROUNDED RISK MITIGATION ====================
ws11 = wb.create_sheet(title="Matriks_Mitigasi_Risiko")
ws11.views.sheetView[0].showGridLines = True

ws11.merge_cells('A1:D1')
ws11['A1'] = "MATRIKS GROUNDED: ACCOMPLISHMENT, ISSUES & RISK MITIGATION (9 SISI)"
ws11['A1'].font = font_title
ws11['A1'].fill = fill_navy_header
ws11['A1'].alignment = align_center
ws11.row_dimensions[1].height = 24

headers11 = ["Perspektif BSC", "Accomplishment (Pencapaian Teruji)", "Issues & Root Cause (Tantangan)", "Next Steps & Risk Mitigation (Solusi)"]
for col_idx, h in enumerate(headers11, 1):
    cell = ws11.cell(row=3, column=col_idx)
    cell.value = h
    cell.font = font_header
    cell.fill = fill_orange_header
    cell.alignment = align_center
    cell.border = thin_border
ws11.row_dimensions[3].height = 22

matrix_rows11 = [
    ("1. Apps / Platform", "Lolos Apple Review 5.1.1 & Guest Browsing aktif 100%.", "Kios konvensional butuh adaptasi alur digital.", "Rilis fitur Quick Re-Order 1-klik & panduan video tutorial."),
    ("2. Customer", "Pemesanan mandiri 24/7 & live chat CS terintegrasi.", "Kios daerah terbiasa pesan via telepon pribadi sales.", "Insentif kupon promo belanja perdana via aplikasi mobile."),
    ("3. Sales", "Penugasan toko binaan terkunci rapi per salesman_id.", "Kekhawatiran pemotongan komisi pesanan mobile.", "SK Direksi: komisi 100% tetap milik sales pembina kios."),
    ("4. Logistik", "Ongkir presisi kecamatan & quote lock 30 menit teruji.", "Tarif kurir reguler mahal untuk pupuk tonase besar.", "Aktivasi opsi armada truk internal Karisma rute terjadwal."),
    ("5. Multiplatform", "Tri-Platform iOS Swift, Android Kotlin & Web Admin siap.", "Variasi spesifikasi ponsel Android mitra di daerah.", "Menjaga ukuran APK < 15MB & kompresi foto lokal perangkat."),
    ("6. Perusahaan", "Settlement kas cepat memotong DSO; audit faktur aman.", "Penyesuaian kebiasaan staf cabang ke sistem otomatis.", "Pelatihan SOP cabang & monitoring omset terpusat."),
    ("7. Keuangan", "BRIVA auto-settlement & margin 3 level terkunci.", "Risiko lonjakan antrean saat puncak musim tanam.", "Idempotency random external-id & fallback transfer kasir."),
    ("8. Teknis", "Backend RESTful API /api/v1 & transaksi ACID DB.", "Volume data transaksi historis semakin membesar.", "Penjadwalan archiving data berkala & optimasi indeks."),
    ("9. Keamanan", "Bearer Token SHA-256, BCRYPT, audit non-PII lulus uji.", "Kebutuhan rotasi kunci privat enkripsi berkala.", "Prosedur rotasi kunci tahunan & monitoring log server.")
]

for idx, (mw, acc, isu, sol) in enumerate(matrix_rows11, 4):
    ws11.cell(row=idx, column=1, value=mw).font = font_bold
    ws11.cell(row=idx, column=2, value=acc).font = font_regular
    ws11.cell(row=idx, column=3, value=isu).font = font_regular
    ws11.cell(row=idx, column=4, value=sol).font = font_regular
    
    for c in range(1, 5):
        ws11.cell(row=idx, column=c).border = thin_border
    ws11.row_dimensions[idx].height = 24

ws11.column_dimensions['A'].width = 24
ws11.column_dimensions['B'].width = 40
ws11.column_dimensions['C'].width = 40
ws11.column_dimensions['D'].width = 45

excel_path = "/Applications/XAMPP/xamppfiles/htdocs/kiustore/docs/nw_bsc_scorecard_karisma_online.xlsx"
wb.save(excel_path)
print(f"SUCCESS: Generated 9 BSC Excel Scorecard at {excel_path}")
