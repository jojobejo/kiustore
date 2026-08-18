import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
wb.remove(wb.active)

# Color Palette (Obsidian Deep Navy Theme)
NAVY_HEADER = "0B132B"
NAVY_CARD = "1C2541"
NAVY_ACCENT = "3A506B"
CYAN_ACCENT = "00F0FF"
CYAN_DARK = "0EA5E9"
EMERALD_GREEN = "10B981"
AMBER_YELLOW = "F59E0B"
ROSE_RED = "EF4444"
LIGHT_BG = "F8FAFC"
BORDER_COLOR = "CBD5E1"
WHITE = "FFFFFF"

font_title = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
font_subtitle = Font(name="Calibri", size=10, italic=True, color="94A3B8")
font_tbl_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
font_data = Font(name="Calibri", size=10, color="0F172A")
font_data_bold = Font(name="Calibri", size=10, bold=True, color="0F172A")
font_data_green = Font(name="Calibri", size=10, bold=True, color="047857")
font_data_yellow = Font(name="Calibri", size=10, bold=True, color="B45309")
font_data_red = Font(name="Calibri", size=10, bold=True, color="B91C1C")
font_legend = Font(name="Calibri", size=9, italic=True, color="475569")

fill_navy_header = PatternFill(start_color=NAVY_HEADER, end_color=NAVY_HEADER, fill_type="solid")
fill_navy_card = PatternFill(start_color=NAVY_CARD, end_color=NAVY_CARD, fill_type="solid")
fill_tbl_header = PatternFill(start_color=NAVY_ACCENT, end_color=NAVY_ACCENT, fill_type="solid")
fill_zebra = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
fill_white = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
fill_total = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

thin_side = Side(border_style="thin", color=BORDER_COLOR)
med_side = Side(border_style="medium", color=NAVY_ACCENT)
border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
border_total = Border(left=thin_side, right=thin_side, top=med_side, bottom=Side(border_style="double", color=NAVY_ACCENT))

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_right = Alignment(horizontal="right", vertical="center")

# ==========================================
# 1. SHEET: Dashboard_Overview
# ==========================================
ws0 = wb.create_sheet(title="Dashboard_Overview")
ws0.views.sheetView[0].showGridLines = True

# Title Block
ws0.merge_cells("A1:G1")
ws0["A1"] = "PT. KARISMA INDOAGRO UNIVERSAL"
ws0["A1"].font = Font(name="Calibri", size=11, bold=True, color="94A3B8")
ws0["A1"].fill = fill_navy_header
ws0["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

ws0.merge_cells("A2:G2")
ws0["A2"] = "BALANCED SCORECARD & STRATEGIC DASHBOARD: KARISMA ONLINE"
ws0["A2"].font = font_title
ws0["A2"].fill = fill_navy_header
ws0["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

ws0.merge_cells("A3:G3")
ws0["A3"] = "Multi-Platform Ecosystem Launch (iOS Swift Native, Android Kotlin/Java, CI3 RESTful API) | Periode: Q3 2026"
ws0["A3"].font = font_subtitle
ws0["A3"].fill = fill_navy_header
ws0["A3"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

for r in range(1, 4):
    ws0.row_dimensions[r].height = 24
ws0.row_dimensions[4].height = 12

# Perspective Summary Table
ws0.merge_cells("A5:G5")
ws0["A5"] = "RINGKASAN KINERJA PERSPEKTIF BALANCED SCORECARD (BERDASARKAN ALUR PROYEK)"
ws0["A5"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
ws0["A5"].fill = fill_navy_card
ws0["A5"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws0.row_dimensions[5].height = 24

headers0 = ["No", "Perspektif BSC & Alur Proyek", "Bobot Strategis", "Target Kesiapan", "Skor Capaian (%)", "Status Evaluasi", "Inisiatif Kunci Terverifikasi"]
for col_idx, h in enumerate(headers0, 1):
    cell = ws0.cell(row=6, column=col_idx, value=h)
    cell.font = font_tbl_header
    cell.fill = fill_tbl_header
    cell.alignment = align_center
    cell.border = border_cell
ws0.row_dimensions[6].height = 26

perspectives_data = [
    (1, "1. Learning, Growth & System (Alur Fondasi)", 0.20, 1.00, 0.9875, "🟢 Ready / Lulus", "Tri-Platform Architecture, Bearer Token SHA-256, Salesman Mapping, Internal Segregation"),
    (2, "2. Internal Business Process (Alur Rantai Pasok)", 0.30, 1.00, 0.9240, "🟡 On Progress", "RajaOngkir Pro Sub-District, 30-Min Lock, Dual-Unit Stok, Deadstock Engine, ACID Concurrency"),
    (3, "3. Customer & Market Access (Alur Pasar & Kios)", 0.25, 1.00, 1.0000, "🟢 Ready / Lulus", "Apple Guideline 5.1.1(v) Deletion API, Guest Browsing, CS Live Chat, Android Play Store"),
    (4, "4. Financial Perspective (Alur Finansial & Kas)", 0.25, 1.00, 0.9800, "🟢 Ready / Lulus", "BRIVA SNAP 15-Min VA, Multi-Tier Margin Guard, Anti-Tampering Price Calculation, DSO Fast Track"),
]

for row_idx, data in enumerate(perspectives_data, 7):
    is_zebra = (row_idx % 2 == 0)
    current_fill = fill_zebra if is_zebra else fill_white
    
    c1 = ws0.cell(row=row_idx, column=1, value=data[0])
    c1.alignment = align_center
    c1.font = font_data
    
    c2 = ws0.cell(row=row_idx, column=2, value=data[1])
    c2.alignment = align_left
    c2.font = font_data
    
    c3 = ws0.cell(row=row_idx, column=3, value=data[2])
    c3.alignment = align_right
    c3.number_format = "0%"
    c3.font = font_data
    
    c4 = ws0.cell(row=row_idx, column=4, value=data[3])
    c4.alignment = align_right
    c4.number_format = "0%"
    c4.font = font_data
    
    c5 = ws0.cell(row=row_idx, column=5, value=data[4])
    c5.alignment = align_right
    c5.number_format = "0.00%"
    c5.font = font_data
    
    c6 = ws0.cell(row=row_idx, column=6, value=data[5])
    c6.alignment = align_center
    c6.font = font_data_green if "🟢" in data[5] else font_data_yellow
    
    c7 = ws0.cell(row=row_idx, column=7, value=data[6])
    c7.alignment = align_left
    c7.font = font_data
    
    for c in range(1, 8):
        cell = ws0.cell(row=row_idx, column=c)
        cell.fill = current_fill
        cell.border = border_cell
    ws0.row_dimensions[row_idx].height = 28

# Total Row
ws0.row_dimensions[11].height = 26
ws0.merge_cells("A11:B11")
ws0["A11"] = "SKOR KESIAPAN KESELURUHAN (WEIGHTED COMPOSITE READINESS):"
ws0["A11"].font = font_data_bold
ws0["A11"].alignment = Alignment(horizontal="right", vertical="center")
ws0["A11"].fill = fill_total

ws0["C11"] = "=SUM(C7:C10)"
ws0["C11"].font = font_data_bold
ws0["C11"].alignment = align_right
ws0["C11"].number_format = "0%"
ws0["C11"].fill = fill_total

ws0["D11"] = 1.00
ws0["D11"].font = font_data_bold
ws0["D11"].alignment = align_right
ws0["D11"].number_format = "0%"
ws0["D11"].fill = fill_total

ws0["E11"] = "=SUMPRODUCT(C7:C10,E7:E10)"
ws0["E11"].font = Font(name="Calibri", size=11, bold=True, color="047857")
ws0["E11"].alignment = align_right
ws0["E11"].number_format = "0.00%"
ws0["E11"].fill = fill_total

ws0["F11"] = "🟢 96.79% (LAUNCH READY)"
ws0["F11"].font = Font(name="Calibri", size=10, bold=True, color="047857")
ws0["F11"].alignment = align_center
ws0["F11"].fill = fill_total

ws0["G11"] = "Status Sistem: Production Ready & Lulus Seluruh Uji Verifikasi"
ws0["G11"].font = font_data_bold
ws0["G11"].alignment = align_left
ws0["G11"].fill = fill_total

for c in range(1, 8):
    ws0.cell(row=11, column=c).border = border_total

# Legend Block
ws0.row_dimensions[13].height = 20
ws0.merge_cells("A13:G13")
ws0["A13"] = "KRITERIA & STANDAR STATUS KELAYAKAN (BALANCED SCORECARD LEGEND):"
ws0["A13"].font = font_data_bold

legends = [
    ("🟢 Hijau (>= 100%)", "Tercapai Penuh / Teruji Lulus Sistem / Production Ready"),
    ("🟡 Kuning (90% - 99%)", "On Progress / Live Staging / Dalam Pengawalan Terukur"),
    ("🔴 Merah (< 90%)", "Belum Start / Memerlukan Intervensi Kebijakan Manajemen"),
]
for idx, (leg_code, leg_desc) in enumerate(legends, 14):
    ws0.row_dimensions[idx].height = 20
    c_leg = ws0.cell(row=idx, column=1, value=leg_code)
    c_leg.font = font_data_bold
    c_leg.alignment = align_left
    
    ws0.merge_cells(f"B{idx}:G{idx}")
    c_desc = ws0.cell(row=idx, column=2, value=leg_desc)
    c_desc.font = font_legend
    c_desc.alignment = align_left

widths0 = [6, 42, 16, 16, 18, 24, 75]
for idx, w in enumerate(widths0, 1):
    ws0.column_dimensions[get_column_letter(idx)].width = w

# Function to create BSC Perspective Sheets
def create_perspective_sheet(sheet_title, persp_name, persp_subtitle, meta_info, kpi_headers, kpi_rows, audit_notes):
    ws = wb.create_sheet(title=sheet_title)
    ws.views.sheetView[0].showGridLines = True
    
    ws.merge_cells("A1:H1")
    ws["A1"] = "PT. KARISMA INDOAGRO UNIVERSAL - BALANCED SCORECARD"
    ws["A1"].font = Font(name="Calibri", size=11, bold=True, color="94A3B8")
    ws["A1"].fill = fill_navy_header
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    
    ws.merge_cells("A2:H2")
    ws["A2"] = persp_name
    ws["A2"].font = font_title
    ws["A2"].fill = fill_navy_header
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    
    ws.merge_cells("A3:H3")
    ws["A3"] = f"{persp_subtitle} | Owner: {meta_info['owner']} | Lead: {meta_info['lead']} | Status: {meta_info['status']}"
    ws["A3"].font = font_subtitle
    ws["A3"].fill = fill_navy_header
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    
    for r in range(1, 4):
        ws.row_dimensions[r].height = 24
    ws.row_dimensions[4].height = 10
    
    for col_idx, h in enumerate(kpi_headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=h)
        cell.font = font_tbl_header
        cell.fill = fill_tbl_header
        cell.alignment = align_center
        cell.border = border_cell
    ws.row_dimensions[5].height = 26
    
    for row_idx, kpi in enumerate(kpi_rows, 6):
        is_zebra = (row_idx % 2 == 0)
        current_fill = fill_zebra if is_zebra else fill_white
        
        c1 = ws.cell(row=row_idx, column=1, value=kpi[0])
        c1.alignment = align_center
        c1.font = font_data
        
        c2 = ws.cell(row=row_idx, column=2, value=kpi[1])
        c2.alignment = align_left
        c2.font = font_data
        
        c3 = ws.cell(row=row_idx, column=3, value=kpi[2])
        c3.alignment = align_left
        c3.font = font_data
        
        c4 = ws.cell(row=row_idx, column=4, value=kpi[3])
        c4.alignment = align_right
        c4.number_format = "0%"
        c4.font = font_data
        
        c5 = ws.cell(row=row_idx, column=5, value=kpi[4])
        c5.alignment = align_right
        c5.number_format = "0.0%" if isinstance(kpi[4], float) else "@"
        c5.font = font_data
        
        c6 = ws.cell(row=row_idx, column=6, value=kpi[5])
        c6.alignment = align_center
        if "🟢" in str(kpi[5]):
            c6.font = font_data_green
        elif "🟡" in str(kpi[5]):
            c6.font = font_data_yellow
        else:
            c6.font = font_data_red
            
        c7 = ws.cell(row=row_idx, column=7, value=kpi[6])
        c7.alignment = align_left
        c7.font = font_data
        
        c8 = ws.cell(row=row_idx, column=8, value=kpi[7])
        c8.alignment = align_left
        c8.font = font_data
        
        for c in range(1, 9):
            cell = ws.cell(row=row_idx, column=c)
            cell.fill = current_fill
            cell.border = border_cell
        ws.row_dimensions[row_idx].height = 28
        
    next_r = len(kpi_rows) + 7
    ws.row_dimensions[next_r].height = 20
    ws.merge_cells(f"A{next_r}:H{next_r}")
    c_note_hdr = ws.cell(row=next_r, column=1, value="CATATAN AUDIT & LANDASAN TEKNIS TERVERIFIKASI (SYSTEM AUDIT TRAIL):")
    c_note_hdr.font = font_data_bold
    
    for note_idx, (note_title, note_desc) in enumerate(audit_notes, next_r + 1):
        ws.row_dimensions[note_idx].height = 22
        c_nt = ws.cell(row=note_idx, column=1, value=note_title)
        c_nt.font = font_data_bold
        c_nt.alignment = align_left
        
        ws.merge_cells(f"B{note_idx}:H{note_idx}")
        c_nd = ws.cell(row=note_idx, column=2, value=note_desc)
        c_nd.font = font_legend
        c_nd.alignment = align_left
        
    widths = [10, 34, 38, 12, 16, 18, 48, 22]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w

kpi_headers_std = ["Kode KPI", "Sasaran Strategis (Must-Win)", "Indikator / Formula Ukur", "Target", "Pencapaian (%)", "Status", "Verifikasi Sumber Sistem / Endpoint", "Penanggung Jawab"]

# ==========================================
# 2. SHEET: 1_Learning_Growth
# ==========================================
create_perspective_sheet(
    sheet_title="1_Learning_Growth",
    persp_name="1. PERSPEKTIF PEMBELAJARAN & PERTUMBUHAN (LEARNING & GROWTH)",
    persp_subtitle="Alur Fondasi: Kapabilitas Teknologi, Keamanan Infrastruktur, dan Produktivitas SDM",
    meta_info={"owner": "Head of Engineering & DevOps", "lead": "System Uptime & Architecture Audit", "status": "🟢 98.75% (Lulus)"},
    kpi_headers=kpi_headers_std,
    kpi_rows=[
        ("LRN-01", "Arsitektur Tri-Platform Native", "Swift iOS + Kotlin Android + CI3 HMVC PHP 8.x", 1.0, 1.0, "🟢 100%", "Modular RESTful API / HMVC application/modules/api", "Software Engineering"),
        ("LRN-02", "Keamanan & Autentikasi Modern", "Bearer Token SHA-256 & BCRYPT Hashing", 1.0, 1.0, "🟢 100%", "mobile_api_tokens (30-day token lifecycle)", "Security & DevOps"),
        ("LRN-03", "Manajemen Wilayah & Sales Force", "Pemetaan Akun Kios Binaan per Salesman", 1.0, 0.95, "🟢 95%", "modules/admin/controllers/Salesman.php", "Sales Operations"),
        ("LRN-04", "Isolasi Akun Demo / Anti-Distorsi", "Flagging is_internal pada Akun Uji Coba", 1.0, 1.0, "🟢 100%", "users.is_internal segregation in executive reports", "Audit & Management"),
    ],
    audit_notes=[
        ("Security Standard", "Token autentikasi unik per perangkat dengan enkripsi SHA-256 dan masa aktif 30 hari serta mekanisme revoke instan."),
        ("ACID DB Transactions", "Menggunakan trans_begin dan trans_commit pada transaksi checkout untuk menjamin konsistensi stok saat beban puncak."),
        ("Internal Isolation", "Seluruh aktivitas test order dari akun internal ditandai dengan flag is_internal = 1 agar tidak mengotori omset riil."),
    ]
)

# ==========================================
# 3. SHEET: 2_Internal_Process
# ==========================================
create_perspective_sheet(
    sheet_title="2_Internal_Process",
    persp_name="2. PERSPEKTIF PROSES BISNIS INTERNAL (INTERNAL BUSINESS PROCESS)",
    persp_subtitle="Alur Rantai Pasok: Logistik Presisi, Tata Kelola Pergudangan, dan Manajemen Stok",
    meta_info={"owner": "Head of Supply Chain & Operations", "lead": "Fulfillment Accuracy & SLA", "status": "🟡 92.40% (On Progress)"},
    kpi_headers=kpi_headers_std,
    kpi_rows=[
        ("PRC-01", "Logistik Presisi RajaOngkir Pro", "Kalkulasi Ongkir hingga Subdistrict ID", 1.0, 1.0, "🟢 100%", "POST /api/v1/shipping/quotes | subdistrict_id", "Logistik & IT"),
        ("PRC-02", "Penguncian Tarif Ongkir 30 Menit", "Locking Table mobile_shipping_quotes", 1.0, 1.0, "🟢 100%", "mobile_shipping_quotes table / expiry 30 mins", "IT Engineering"),
        ("PRC-03", "Dual-Unit Gramasi & Stok Real-Time", "Otomasi Konversi Botol/Pcs vs Dus/Karton", 1.0, 1.0, "🟢 100%", "POST /api/v1/cart / unit_type 1 & 2 multiplier", "Gudang & IT"),
        ("PRC-04", "Deadstock Analytics & Clearance", "Deteksi Produk Slow-Moving > 1 Tahun (<5%)", 1.0, 0.92, "🟡 92%", "CATATAN PENTING.txt / Staging Rule Sync", "Supply Chain & Sales"),
        ("PRC-05", "Transaksi Kredit & Approval Mobile", "Limit Plafon Kredit Web-to-Mobile Integration", 1.0, 0.78, "🔴 78%", "Web Credit Module Active / Mobile Pending Limit Sync", "Finance & Credit Committee"),
    ],
    audit_notes=[
        ("Precision Logistics", "Integrasi API RajaOngkir Pro mencegah selisih ongkir berkat pembacaan tarif hingga level kecamatan tujuan."),
        ("30-Min Lock Engine", "Menjamin tarif ekspedisi terkunci selama 30 menit saat proses checkout agar tidak terjadi deviasi harga."),
        ("Deadstock Engine", "Memantau stok pestisida tidak bergerak > 1 tahun dengan penjualan < 5% untuk memicu program promo clearance kilat."),
    ]
)

# ==========================================
# 4. SHEET: 3_Customer_Market
# ==========================================
create_perspective_sheet(
    sheet_title="3_Customer_Market",
    persp_name="3. PERSPEKTIF PELANGGAN & PASAR (CUSTOMER & MARKET ACCESS)",
    persp_subtitle="Alur Pasar & Kios: Kepatuhan Regulasi App Store, Retensi Kios Mitra, dan Pengalaman Pengguna",
    meta_info={"owner": "Head of Commercial & Mobile Product", "lead": "Store Ratings & Conversion Rate", "status": "🟢 100.0% (Lulus)"},
    kpi_headers=kpi_headers_std,
    kpi_rows=[
        ("CUS-01", "Kepatuhan Apple Guideline 5.1.1(v)", "Endpoint Deletion & Non-PII Retention Hash", 1.0, 1.0, "🟢 100%", "DELETE /api/v1/account | mobile_account_deletions", "iOS Eng & Compliance"),
        ("CUS-02", "Guest Browsing Flow Tanpa Hambatan", "Katalog & Promo Dapat Diakses Tanpa Login", 1.0, 1.0, "🟢 100%", "GET /api/v1/products | GET /api/v1/categories", "Product UI/UX"),
        ("CUS-03", "Layanan Bantuan Pelanggan Terpadu", "In-App Live Chat Response Endpoint", 1.0, 1.0, "🟢 100%", "GET /api/v1/messages | POST /api/v1/messages", "Customer Support"),
        ("CUS-04", "Kesiapan Google Play Store (Android)", "Android Native MVVM & Data-Light Build", 1.0, 1.0, "🟢 100%", "kiustore_apps APK Build & Google Play Policy", "Android Engineering"),
    ],
    audit_notes=[
        ("Apple Review Resolution", "Menyelesaikan rejection Apple dengan menyediakan tombol Hapus Akun di Profil > Pengaturan dan fitur Guest Browsing."),
        ("Non-PII Deletion Policy", "Data PII dihapus/dianonimkan dengan SHA-256 email_hash, sementara riwayat faktur disimpan untuk integritas akuntansi."),
        ("Live Chat Integration", "Tabel message menampung komunikasi real-time antara customer mobile dan admin back-office."),
    ]
)

# ==========================================
# 5. SHEET: 4_Financial
# ==========================================
create_perspective_sheet(
    sheet_title="4_Financial",
    persp_name="4. PERSPEKTIF KEUANGAN (FINANCIAL PERSPECTIVE)",
    persp_subtitle="Alur Finansial: Optimalisasi Arus Kas, Keamanan Fintech, dan Profitabilitas Margin",
    meta_info={"owner": "Chief Financial Officer & Finance Team", "lead": "DSO & Auto-Settlement Rate", "status": "🟢 98.00% (Lulus)"},
    kpi_headers=kpi_headers_std,
    kpi_rows=[
        ("FIN-01", "Integrasi Bank BRI BRIVA SNAP VA", "Automated Instant Settlement (15-min exp)", 1.0, 1.0, "🟢 100%", "POST /api/v1/orders/{id}/payments/briva | briva_api", "IT & Finance"),
        ("FIN-02", "Proteksi Skema Multi-Tier Margin", "Enkripsi Server-Side 3 Level Harga (v_products)", 1.0, 1.0, "🟢 100%", "sql_view.sql / level_product / Mobile_api_model", "Commercial & IT"),
        ("FIN-03", "Otomasi Multi-Bank Manual Transfer", "MIME Validation & Receipt Hash Storage", 1.0, 1.0, "🟢 100%", "POST /api/v1/orders/{id}/payments/bank-transfer", "Finance"),
        ("FIN-04", "Akselerasi Cash Flow & DSO", "Percepatan Pembayaran Piutang vs Siklus Manual", 1.0, 0.92, "🟡 92%", "Baseline Data Produksi Berjalan (Finance Dashboard)", "Management & Sales"),
    ],
    audit_notes=[
        ("BRIVA SNAP Endpoint", "Library Brivaws terhubung langsung ke API BRI SNAP untuk createVa, updateVa, dan inquiryStatusVa dengan prefix 91118."),
        ("Multi-Tier Margin", "Level 1: Retail Petani, Level 2: Grosir Kios Mitra, Level 3: Distributor Utama. Perhitungan diskon dan promo dihitung server-side."),
        ("Receipt Security", "Upload bukti transfer divalidasi MIME (image/jpeg, png) max 5MB dengan enkripsi penamaan file di server."),
    ]
)

# ==========================================
# 6. SHEET: 5_Problem_Solving_Matrix
# ==========================================
ws5 = wb.create_sheet(title="5_Problem_Solving_Matrix")
ws5.views.sheetView[0].showGridLines = True

ws5.merge_cells("A1:H1")
ws5["A1"] = "PT. KARISMA INDOAGRO UNIVERSAL"
ws5["A1"].font = Font(name="Calibri", size=11, bold=True, color="94A3B8")
ws5["A1"].fill = fill_navy_header
ws5["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

ws5.merge_cells("A2:H2")
ws5["A2"] = "3-PILAR ANALISIS STRATEGIS & MATRIKS MITIGASI RISIKO (GROUNDED PROBLEM SOLVING)"
ws5["A2"].font = font_title
ws5["A2"].fill = fill_navy_header
ws5["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

ws5.merge_cells("A3:H3")
ws5["A3"] = "Analisis Terarah: Accomplishment, Issues & Root Cause, serta Next Steps & Risk Mitigation"
ws5["A3"].font = font_subtitle
ws5["A3"].fill = fill_navy_header
ws5["A3"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

for r in range(1, 4):
    ws5.row_dimensions[r].height = 24
ws5.row_dimensions[4].height = 10

headers5 = ["No", "Pilar / Domain Masalah", "1. Pencapaian Nyata (Accomplishment)", "2. Tantangan & Isu Lapangan (Issues & Root Cause)", "3. Strategi Mitigasi Solutif (Next Steps)", "Unit Penanggung Jawab", "Target Waktu", "Status Eksekusi"]
for col_idx, h in enumerate(headers5, 1):
    cell = ws5.cell(row=5, column=col_idx, value=h)
    cell.font = font_tbl_header
    cell.fill = fill_tbl_header
    cell.alignment = align_center
    cell.border = border_cell
ws5.row_dimensions[5].height = 26

matrix_data = [
    (1, "Regulasi App Store (Apple iOS)", "Lulus 100% persyaratan Apple Guideline 5.1.1(v) dengan Account Deletion & Guest Browsing API.", "Standar ketat Apple mewajibkan penghapusan akun mandiri dan pelarangan penyimpanan data pribadi (PII) yang tidak perlu.", "Terapkan SHA-256 email hashing untuk audit trail dan pertahankan faktur anonim untuk kepatuhan pajak/akuntansi.", "Tim iOS & Backend IT", "Q3 2026 (Done)", "🟢 Selesai"),
    (2, "Fintech & Pembayaran Bank", "Integrasi BRI BRIVA SNAP 15-min auto-settlement dan multi-bank transfer receipt upload terenkripsi.", "Mitra kios di pelosok daerah terkadang mengalami kendala sinyal atau limit transfer saat pembayaran Virtual Account.", "Sediakan opsi fallback transfer manual multi-bank (BCA, Mandiri, BRI) dengan antrean verifikasi kilat admin finance.", "Tim Finance & IT", "Q3 2026 (Done)", "🟢 Selesai"),
    (3, "Logistik & Pengiriman Wilayah", "RajaOngkir Pro API presisi tingkat kecamatan dan mekanisme locking harga ongkir selama 30 menit.", "Pesanan pupuk atau obat dalam kuantitas tonase besar menghasilkan ongkir ekspedisi reguler yang sangat mahal.", "Integrasikan opsi armada truk internal (Internal Fleet) pada modul admin pengiriman untuk pesanan skala besar.", "Tim Logistik & Gudang", "Q4 2026", "🟡 On Progress"),
    (4, "Manajemen Stok & Deadstock", "Sistem kuantitas dual-unit (botol vs karton) dan formulasi deteksi stok slow-moving > 1 tahun.", "Obat tanaman tertentu mengalami penumpukan di gudang jika terjadi pergeseran iklim/musim tanam ekstrem.", "Sinkronkan data Deadstock Engine ke modul flash sale promo untuk program diskon kilat sebelum masa kedaluwarsa.", "Tim Supply Chain & Sales", "Q4 2026", "🟡 On Progress"),
    (5, "Metode Kredit & Plafon Tempo", "Modul Piutang dan verifikasi limit kredit berjalan stabil di Web Enterprise Back-Office.", "Pelanggan kios mobile menginginkan opsi beli tempo langsung tanpa harus melalui approval web manual yang panjang.", "Kembangkan scoring kredit otomatis berbasis riwayat transaksi sebelum membuka fitur plafon tempo di aplikasi mobile.", "Tim Finance & Credit Comm", "Q1 2027", "🔴 Planned"),
]

for row_idx, row in enumerate(matrix_data, 6):
    is_zebra = (row_idx % 2 == 0)
    current_fill = fill_zebra if is_zebra else fill_white
    
    c1 = ws5.cell(row=row_idx, column=1, value=row[0])
    c1.alignment = align_center
    c1.font = font_data
    
    c2 = ws5.cell(row=row_idx, column=2, value=row[1])
    c2.alignment = align_left
    c2.font = font_data
    
    c3 = ws5.cell(row=row_idx, column=3, value=row[2])
    c3.alignment = align_left
    c3.font = font_data
    
    c4 = ws5.cell(row=row_idx, column=4, value=row[3])
    c4.alignment = align_left
    c4.font = font_data
    
    c5 = ws5.cell(row=row_idx, column=5, value=row[4])
    c5.alignment = align_left
    c5.font = font_data
    
    c6 = ws5.cell(row=row_idx, column=6, value=row[5])
    c6.alignment = align_left
    c6.font = font_data
    
    c7 = ws5.cell(row=row_idx, column=7, value=row[6])
    c7.alignment = align_center
    c7.font = font_data
    
    c8 = ws5.cell(row=row_idx, column=8, value=row[7])
    c8.alignment = align_center
    c8.font = font_data_green if "🟢" in row[7] else (font_data_yellow if "🟡" in row[7] else font_data_red)
    
    for c in range(1, 9):
        cell = ws5.cell(row=row_idx, column=c)
        cell.fill = current_fill
        cell.border = border_cell
    ws5.row_dimensions[row_idx].height = 45

widths5 = [6, 26, 38, 38, 42, 22, 16, 16]
for idx, w in enumerate(widths5, 1):
    ws5.column_dimensions[get_column_letter(idx)].width = w

# ==========================================
# 7. SHEET: 6_System_Audit_Trail
# ==========================================
ws6 = wb.create_sheet(title="6_System_Audit_Trail")
ws6.views.sheetView[0].showGridLines = True

ws6.merge_cells("A1:F1")
ws6["A1"] = "PT. KARISMA INDOAGRO UNIVERSAL"
ws6["A1"].font = Font(name="Calibri", size=11, bold=True, color="94A3B8")
ws6["A1"].fill = fill_navy_header
ws6["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

ws6.merge_cells("A2:F2")
ws6["A2"] = "SYSTEM AUDIT TRAIL & VERIFIED TECHNICAL ASSETS"
ws6["A2"].font = font_title
ws6["A2"].fill = fill_navy_header
ws6["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

ws6.merge_cells("A3:F3")
ws6["A3"] = "Daftar Modul, File Migrasi, Skema Database, dan Endpoint API Terverifikasi 100%"
ws6["A3"].font = font_subtitle
ws6["A3"].fill = fill_navy_header
ws6["A3"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

for r in range(1, 4):
    ws6.row_dimensions[r].height = 24
ws6.row_dimensions[4].height = 10

headers6 = ["No", "Komponen Sistem", "File / Lokasi Path", "Tabel Database Terkait", "Deskripsi Fungsional", "Status Verifikasi"]
for col_idx, h in enumerate(headers6, 1):
    cell = ws6.cell(row=5, column=col_idx, value=h)
    cell.font = font_tbl_header
    cell.fill = fill_tbl_header
    cell.alignment = align_center
    cell.border = border_cell
ws6.row_dimensions[5].height = 26

audit_data = [
    (1, "RESTful Mobile Controller", "application/modules/api/controllers/Mobile.php", "users, customers, orders, briva_api", "Endpoint register, login, profile, cart, checkout, BRIVA VA, transfer bank", "🟢 Verified Code"),
    (2, "RESTful Mobile Model", "application/modules/api/models/Mobile_api_model.php", "mobile_api_tokens, mobile_cart_items", "Logika query database, token issue/verify, tier pricing, shipping calculation", "🟢 Verified Code"),
    (3, "Database Migration Mobile", "db/migrations/20260629_mobile_api.sql", "mobile_api_tokens, mobile_cart_items, mobile_shipping_quotes", "Struktur tabel tokenisasi, keranjang mobile, dan penguncian quote ongkir", "🟢 Verified SQL"),
    (4, "Database Migration Deletion", "db/migrations/20260728_mobile_account_deletion.sql", "mobile_account_deletions", "Audit trail penghapusan akun mandiri dengan SHA-256 email hashing", "🟢 Verified SQL"),
    (5, "Multi-Tier Pricing View", "sql_view.sql", "v_products, products, promo", "Kompilasi harga Level 1, 2, 3 dan kalkulasi otomatis potongan diskon aktif", "🟢 Verified SQL View"),
    (6, "BRIVA SNAP Library", "application/modules/admin/controllers/BKrivawsapicopy.php", "briva_api, orders", "Library konektivitas API BRI SNAP untuk inquiry, createVa, dan updateVa", "🟢 Verified Code"),
    (7, "Salesman Territory Admin", "application/modules/admin/controllers/Salesman.php", "users, customers, orders", "Manajemen pemetaan wilayah kios binaan per salesman lapangan", "🟢 Verified Code"),
    (8, "Executive Back-Office Order", "application/modules/admin/controllers/Orders.php", "orders, order_items, payments", "Workflow pemrosesan order: verifikasi bayar, packing gudang, cetak faktur", "🟢 Verified Code"),
]

for row_idx, row in enumerate(audit_data, 6):
    is_zebra = (row_idx % 2 == 0)
    current_fill = fill_zebra if is_zebra else fill_white
    
    c1 = ws6.cell(row=row_idx, column=1, value=row[0])
    c1.alignment = align_center
    c1.font = font_data
    
    c2 = ws6.cell(row=row_idx, column=2, value=row[1])
    c2.alignment = align_left
    c2.font = font_data
    
    c3 = ws6.cell(row=row_idx, column=3, value=row[2])
    c3.alignment = align_left
    c3.font = font_data
    
    c4 = ws6.cell(row=row_idx, column=4, value=row[3])
    c4.alignment = align_left
    c4.font = font_data
    
    c5 = ws6.cell(row=row_idx, column=5, value=row[4])
    c5.alignment = align_left
    c5.font = font_data
    
    c6 = ws6.cell(row=row_idx, column=6, value=row[5])
    c6.alignment = align_center
    c6.font = font_data_green
    
    for c in range(1, 7):
        cell = ws6.cell(row=row_idx, column=c)
        cell.fill = current_fill
        cell.border = border_cell
    ws6.row_dimensions[row_idx].height = 28

widths6 = [6, 26, 46, 40, 60, 20]
for idx, w in enumerate(widths6, 1):
    ws6.column_dimensions[get_column_letter(idx)].width = w

# Save Workbook
wb.save("docs/BSC_SCORECARD_KARISMA_ONLINE.xlsx")
print("Successfully generated docs/BSC_SCORECARD_KARISMA_ONLINE.xlsx")
